from openpyxl import load_workbook
import numpy as np
from openpyxl.utils import get_column_letter
import openpyxl
from pathlib import Path
from collections import defaultdict
import pandas as pd
from scipy import stats
import os, sys, re, json, time, shutil, argparse
import sys, os, time, argparse, shutil, subprocess, warnings, importlib.util
import requests, pandas as pd
from Bio.PDB.SASA import ShrakeRupley
from scipy.spatial.distance import cdist
from openpyxl.styles import PatternFill, Font, Alignment
from Bio.PDB import PDBParser, is_aa

"""
糖蛋白 Re-Glyco Ensemble 完整流程脚本
==============================================
用法:
    python glycan_pipeline.py [步骤]

步骤:
    1  verify_ac      从 GlycoShape 验证/查找糖链 AC (GlyTouCan ID)
    2  update_excel   将 AC 验证结果写入 Excel (GlycoShape_ID / 验证状态)
    3  run_ensemble   批量 Re-Glyco Ensemble 建模并下载 PDB
    4  model_status   更新 Excel 中的建模状态列
    5  flatten        整理输出：重命名 PDB、合并 ensemble_stats 到 xlsx
    all              依次执行全部步骤

数据文件 (需放在同一基础目录):
    基质蛋白三物种糖链汇总.xlsx      — 原始表格 (步骤1/2 输入)
    基质蛋白三物种糖链汇总_更新.xlsx  — 步骤2 输出 / 步骤3/4 输入
    glycan_ac_result.json         — 步骤1 输出 / 步骤3 输入
    ReGlyco_Ensemble/             — 步骤3/4/5 输出目录

GlycoShape API: https://glycoshape.org
"""

sys.stdout.reconfigure(encoding="utf-8")

# ══════════════════════════════════════════════════════════════
#  全局配置
# ══════════════════════════════════════════════════════════════
BASE_DIR      = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
API           = "https://glycoshape.org"
XLSX_ORIG     = os.path.join(BASE_DIR, "基质蛋白三物种糖链汇总.xlsx")
XLSX_UPDATED  = os.path.join(BASE_DIR, "基质蛋白三物种糖链汇总_更新.xlsx")
XLSX_MODELED  = os.path.join(BASE_DIR, "基质蛋白三物种糖链汇总_建模.xlsx")
JSON_AC       = os.path.join(BASE_DIR, "glycan_ac_result.json")
OUTDIR        = os.path.dirname(os.path.abspath(__file__))  # ReGlyco_Ensemble/

ENSEMBLE_SIZE = 50
POLL_INTERVAL = 15
MAX_POLL      = 120

# 残基质量 (Da)
RESIDUE_MASS  = {
    "HexNAc": 203.0794,
    "Hex":    162.0528,
    "NeuAc":  291.0954,
    "dHex":   146.0579,
    "Pen":    132.0423,
}
WATER        = 18.0106
MASS_TOL     = 0.5

# 颜色
C_GREEN  = PatternFill("solid", fgColor="C6EFCE")
C_YELLOW = PatternFill("solid", fgColor="FFEB9C")
C_RED    = PatternFill("solid", fgColor="FFC7CE")
C_BLUE   = PatternFill("solid", fgColor="D9E1F2")

# ══════════════════════════════════════════════════════════════
#  步骤 1: 验证 AC
# ══════════════════════════════════════════════════════════════
def step1_verify_ac():
    """从 GlycoShape 批量验证/查找糖链的 AC (GlyTouCan ID)，输出 glycan_ac_result.json"""
    print("\n" + "="*60)
    print("步骤 1: 验证 GlycoShape AC")
    print("="*60)

    xl = pd.ExcelFile(XLSX_ORIG)
    df = xl.parse("糖链")
    df.columns = ["idx", "species", "protein_id", "position", "gtype", "composition", "AC", "note"]

    def parse_oxford(s):
        d = {}
        for m in re.finditer(r"([A-Za-z]+)\((\d+)\)", str(s)):
            d[m.group(1)] = int(m.group(2))
        return d

    def expected_mass(comp):
        m = WATER
        for k, v in comp.items():
            m += v * RESIDUE_MASS.get(k, 0)
        return round(m, 2)

    print("获取 GlycoShape 全库...")
    r = requests.post(f"{API}/api/search", json={"search_string": "all"}, timeout=60)
    all_glycans = r.json().get("results", [])
    print(f"  共 {len(all_glycans)} 条")

    cache = {}

    def get_detail(gid):
        if gid not in cache:
            r = requests.get(f"{API}/api/glycan/{gid}", timeout=10)
            cache[gid] = r.json() if r.status_code == 200 else None
        return cache[gid]

    def verify_comp(detail, comp):
        if not detail:
            return False
        dc = detail.get("archetype", {}).get("composition", {})
        if not dc:
            return False
        for k, v in comp.items():
            actual = dc.get(k, dc.get("Neu5Ac", 0) if k == "NeuAc" else 0)
            if actual != v:
                return False
        for ck, cv in dc.items():
            if cv > 0 and ck not in comp:
                if not (ck == "Neu5Ac" and "NeuAc" in comp):
                    return False
        return True

    results = []
    for i, row in df.iterrows():
        comp_str = str(row["composition"]).strip()
        existing = str(row["AC"]).strip()
        comp     = parse_oxford(comp_str)
        target   = expected_mass(comp)

        candidates = [g for g in all_glycans if g.get("mass") and abs(g["mass"] - target) <= MASS_TOL]
        verified   = []
        for c in candidates:
            detail = get_detail(c["ID"])
            if verify_comp(detail, comp):
                arch = detail.get("archetype", {})
                beta = detail.get("beta", {})
                verified.append({
                    "GS_ID":               c["ID"],
                    "glytoucan_archetype": arch.get("glytoucan"),
                    "glytoucan_beta":      beta.get("glytoucan"),
                    "iupac":               (arch.get("iupac") or "")[:120],
                    "mass":                arch.get("mass"),
                })

        status = "✅ " + verified[0]["GS_ID"] if verified else "❌ 未找到"
        print(f"  [{i:02d}] {comp_str:<30} {existing:<12} {status}")
        results.append({"row": i, "composition": comp_str, "existing_AC": existing,
                         "expected_mass": target, "verified": verified})

    with open(JSON_AC, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n✅ 已保存: {JSON_AC}")


# ══════════════════════════════════════════════════════════════
#  步骤 2: 更新 Excel — AC 验证状态
# ══════════════════════════════════════════════════════════════
def step2_update_excel():
    """将 AC 验证结果写入 Excel，新增 GlycoShape_ID 和验证状态两列"""
    print("\n" + "="*60)
    print("步骤 2: 更新 Excel (AC 验证状态)")
    print("="*60)

    with open(JSON_AC, encoding="utf-8") as f:
        results = json.load(f)

    src = XLSX_ORIG if os.path.exists(XLSX_ORIG) else XLSX_UPDATED
    wb  = load_workbook(src)
    ws  = wb["糖链"]

    ws.cell(1, 9, "GlycoShape_ID").font  = Font(bold=True)
    ws.cell(1, 10, "验证状态").font       = Font(bold=True)
    for col in [9, 10]:
        ws.cell(1, col).fill      = C_BLUE
        ws.cell(1, col).alignment = Alignment(horizontal="center")

    for res in results:
        excel_row  = res["row"] + 2
        existing   = res["existing_AC"]
        verified   = res["verified"]
        exact      = [v for v in verified if v.get("glytoucan_beta") == existing]

        if exact:
            gs_str = exact[0]["GS_ID"]
            status = "✅ AC已确认"
            fill   = C_GREEN
        elif verified:
            gs_ids = list(dict.fromkeys(v["GS_ID"] for v in verified))
            gs_str = " / ".join(gs_ids)
            status = f"⚠️ 组成存在({len(verified)}个异构体)"
            fill   = C_YELLOW
        else:
            gs_str = "-"
            status = "❌ GlycoShape未收录"
            fill   = C_RED

        for col, val in ((9, gs_str), (10, status)):
            c = ws.cell(excel_row, col, val)
            c.fill      = fill
            c.alignment = Alignment(horizontal="center" if col == 9 else "left")

        print(f"  行{excel_row}: {res['composition']:<30} {gs_str:<30} {status}")

    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 22

    try:
        wb.save(XLSX_ORIG)
        print(f"\n✅ 已保存: {XLSX_ORIG}")
    except PermissionError:
        wb.save(XLSX_UPDATED)
        print(f"\n⚠️ 原文件被占用，已另存为: {XLSX_UPDATED}")


# ══════════════════════════════════════════════════════════════
#  步骤 3: Re-Glyco Ensemble 建模
# ══════════════════════════════════════════════════════════════
def _create_session(prot_id):
    r = requests.post(f"{API}/api/sessions", json={"protID": prot_id}, timeout=120)
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Session 创建失败: {r.status_code} {r.text[:200]}")
    uuid = r.json()["session_uuid"]
    r2   = requests.get(f"{API}/api/sessions/{uuid}", timeout=60)
    avail = r2.json().get("glycosylation", {}).get("available", [])
    print(f"  Session {uuid}  可用位点={len(avail)}")
    return uuid, avail

def _submit_job(session_uuid, site_key, glytoucan_id):
    r = requests.post(f"{API}/api/jobs", json={
        "session_uuid":   session_uuid,
        "jobType":        "ensemble",
        "selectedGlycans": {site_key: glytoucan_id},
        "parameters":     {"ensembleSize": ENSEMBLE_SIZE, "calculateSASA": False,
                           "seed": 42, "outputFormat": "PDB"}
    }, timeout=60)
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Job 提交失败: {r.status_code} {r.text[:200]}")
    return r.json()["job_uuid"]

def _poll_job(job_uuid):
    for i in range(MAX_POLL):
        r  = requests.get(f"{API}/api/jobs/{job_uuid}", timeout=30)
        d  = r.json()
        st = d.get("status", "unknown")
        print(f"    轮询[{i+1:3d}] {st}    ", end="\r")
        if st in ("completed", "failed", "error"):
            print()
            return d
        time.sleep(POLL_INTERVAL)
    return {"status": "timeout"}

def _download_results(job_d, out_folder):
    want  = {"all.pdb", "ensemble_analysis.png", "ensemble_stats.csv"}
    os.makedirs(out_folder, exist_ok=True)
    for fi in job_d.get("files", []):
        if fi["name"] not in want:
            continue
        url  = f"{API}{fi['url']}" if fi["url"].startswith("/") else fi["url"]
        resp = requests.get(url, timeout=120, stream=True)
        if resp.status_code == 200:
            path = os.path.join(out_folder, fi["name"])
            with open(path, "wb") as f:
                for chunk in resp.iter_content(8192):
                    f.write(chunk)
            print(f"    ✅ {fi['name']}  ({os.path.getsize(path)/1024/1024:.1f} MB)")

def step3_run_ensemble():
    """批量提交 Re-Glyco Ensemble 并下载结果"""
    print("\n" + "="*60)
    print("步骤 3: Re-Glyco Ensemble 建模")
    print("="*60)
    os.makedirs(OUTDIR, exist_ok=True)

    xlsx_in = XLSX_UPDATED if os.path.exists(XLSX_UPDATED) else XLSX_ORIG
    xl = pd.ExcelFile(xlsx_in)
    df = xl.parse("糖链")
    df.columns = ["idx","species","protein_id","position","gtype","composition",
                  "AC","note","GS_ID","status"]
    df["AC"] = df["AC"].astype(str).str.strip()

    # 构建 AC 映射（优先使用 GlycoShape beta ID）
    with open(JSON_AC, encoding="utf-8") as f:
        ac_data = json.load(f)

    gs_ac_map = {}
    for entry in ac_data:
        r     = entry["row"]
        orig  = entry.get("existing_AC", "").strip()
        vlist = entry.get("verified", [])
        if not vlist:
            continue
        exact = any(v.get("glytoucan_beta") == orig for v in vlist)
        if exact:
            gs_ac_map[r] = (orig, orig)
        else:
            first = vlist[0]
            beta  = first.get("glytoucan_beta") or first.get("glytoucan_archetype")
            if beta:
                gs_ac_map[r] = (beta, orig)

    df_valid          = df[df.index.map(lambda i: i in gs_ac_map)].copy()
    df_valid["position"] = df_valid["position"].astype(int)

    print(f"\n共 {len(df_valid)} 个有效任务:")
    for i, row in df_valid.iterrows():
        use, orig = gs_ac_map[i]
        note = f" -> {use}" if use != orig else ""
        print(f"  {row['species']:8} {row['protein_id']:15} 位点{row['position']:4}"
              f"  {row['composition']:<30}  {orig}{note}")

    sessions = {}
    summary  = []

    for prot_id in df_valid["protein_id"].unique():
        print(f"\n{'='*55}\n蛋白: {prot_id}")
        try:
            uuid, avail = _create_session(prot_id)
            avail_map   = {s["residueID"]: s for s in avail}
        except Exception as e:
            print(f"  ❌ Session 失败: {e}")
            continue

        for row_idx, row in df_valid[df_valid["protein_id"] == prot_id].iterrows():
            species  = row["species"]
            position = int(row["position"])
            comp     = row["composition"]
            orig_ac  = row["AC"]
            use_ac, _ = gs_ac_map.get(row_idx, (orig_ac, orig_ac))

            if position not in avail_map:
                print(f"  ⚠️  位点 {position} 不在结构中，跳过")
                summary.append({"species": species, "protein": prot_id, "position": position,
                                 "AC": orig_ac, "status": "位点不存在"})
                continue

            chain       = avail_map[position]["residueChain"]
            site_key    = f"{position}_{chain}"
            folder_name = f"{species}_{prot_id}_{position}_{orig_ac}"
            out_folder  = os.path.join(OUTDIR, folder_name)

            if os.path.exists(os.path.join(out_folder, "all.pdb")):
                print(f"  ⏭️  {folder_name} 已存在，跳过")
                summary.append({"species": species, "protein": prot_id, "position": position,
                                 "AC": orig_ac, "use_ac": use_ac, "status": "已存在"})
                continue

            note = f" (GlycoShape beta: {use_ac})" if use_ac != orig_ac else ""
            print(f"\n  任务: {folder_name}\n  位点键: {site_key}{note}")

            try:
                job_uuid = _submit_job(uuid, site_key, use_ac)
                print(f"  Job: {job_uuid}")
                job_d    = _poll_job(job_uuid)
                status   = job_d.get("status", "unknown")
                if status == "completed":
                    _download_results(job_d, out_folder)
                    summary.append({"species": species, "protein": prot_id, "position": position,
                                     "AC": orig_ac, "use_ac": use_ac, "status": "completed"})
                else:
                    print(f"  ❌ {status}")
                    summary.append({"species": species, "protein": prot_id, "position": position,
                                     "AC": orig_ac, "use_ac": use_ac, "status": f"失败:{status}"})
            except Exception as e:
                print(f"  ❌ 异常: {e}")
                summary.append({"species": species, "protein": prot_id, "position": position,
                                 "AC": orig_ac, "status": f"异常:{e}"})

    with open(os.path.join(OUTDIR, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    done  = sum(1 for s in summary if s["status"] == "completed")
    print(f"\n✅ 完成 {done}/{len(summary)} 个任务  summary -> {OUTDIR}/summary.json")


# ══════════════════════════════════════════════════════════════
#  步骤 4: 更新 Excel — 建模状态
# ══════════════════════════════════════════════════════════════
def step4_model_status():
    """新增 建模状态 / 建模文件 两列到 Excel"""
    print("\n" + "="*60)
    print("步骤 4: 更新 Excel (建模状态)")
    print("="*60)

    built = [d for d in os.listdir(OUTDIR)
             if os.path.isdir(os.path.join(OUTDIR, d))
             and os.path.exists(os.path.join(OUTDIR, d, "all.pdb"))]

    def find_folders(species, prot, pos, ac, gs_str):
        ac   = str(ac).strip()
        gs   = [g.strip() for g in str(gs_str).split("/")
                if g.strip() not in ("", "-", "nan")]
        pfx  = f"{species}_{prot}_{pos}_"
        hits = []
        for d in sorted(built):
            if pfx + ac in d:
                hits.append(d); continue
            for g in gs:
                if g in d:
                    hits.append(d); break
        return hits

    xlsx_in = XLSX_UPDATED if os.path.exists(XLSX_UPDATED) else XLSX_ORIG
    wb = load_workbook(xlsx_in)
    ws = wb["糖链"]

    col_st  = ws.max_column + 1
    col_dir = ws.max_column + 2
    for col, hdr in ((col_st, "建模状态"), (col_dir, "建模文件")):
        c = ws.cell(1, col, hdr)
        c.font      = Font(bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    xl = pd.ExcelFile(xlsx_in)
    df = xl.parse("糖链")
    df.columns = ["idx","species","protein_id","position","gtype","composition",
                  "AC","note","GS_ID","验证状态"]
    df["AC"]    = df["AC"].astype(str).str.strip()
    df["GS_ID"] = df["GS_ID"].astype(str).str.strip()

    for i, row in df.iterrows():
        species = str(row["species"]).strip()
        prot    = str(row["protein_id"]).strip()
        pos     = str(int(row["position"])) if pd.notna(row["position"]) else ""
        ac      = row["AC"]
        gs_str  = row["GS_ID"]
        matched = find_folders(species, prot, pos, ac, gs_str)
        all_gs  = [g.strip() for g in str(gs_str).split("/")
                   if g.strip() not in ("", "-", "nan")]
        ac_ok   = ac not in ("-", "nan", "NaN", "")

        if not matched:
            status = "❌ GlycoShape未收录" if not ac_ok and not all_gs else "❌ 未建模"
            fill   = C_RED
        else:
            expected = len(all_gs) if all_gs else 1
            if len(matched) >= expected:
                status = "✅ 已建模"
                fill   = C_GREEN
            else:
                status = f"⚠️ 部分建模 ({len(matched)}/{expected})"
                fill   = C_YELLOW

        # 建模文件显示 PDB 文件名列表
        pdb_files = []
        for d in matched:
            pdb_path = os.path.join(OUTDIR, d, "all.pdb")
            if os.path.exists(pdb_path):
                pdb_files.append(d + "/all.pdb")

        for col, val in ((col_st, status), (col_dir, "\n".join(pdb_files) or "-")):
            c = ws.cell(i + 2, col, val)
            c.fill      = fill
            c.alignment = Alignment(horizontal="center" if col == col_st else "left",
                                    vertical="center", wrap_text=True)

        print(f"  Row{i:2d} {species:8} {ac:12} -> {status} [{len(matched)}]")

    ws.column_dimensions[get_column_letter(col_st)].width  = 22
    ws.column_dimensions[get_column_letter(col_dir)].width = 50

    try:
        wb.save(XLSX_MODELED)
        print(f"\n✅ 已保存: {XLSX_MODELED}")
    except PermissionError:
        alt = XLSX_MODELED.replace(".xlsx", "_2.xlsx")
        wb.save(alt)
        print(f"\n⚠️ 另存为: {alt}")


# ══════════════════════════════════════════════════════════════
#  步骤 5: 整理输出（重命名 PDB、合并统计）
# ══════════════════════════════════════════════════════════════
def step5_flatten():
    """将 all.pdb 重命名为 {物种}_{Glycan}.pdb，合并 ensemble_stats.csv 到 xlsx"""
    print("\n" + "="*60)
    print("步骤 5: 整理输出文件")
    print("="*60)

    def parse_folder(name):
        parts   = name.split("_")
        species = parts[0]
        pos_idx = next((i for i, p in enumerate(parts) if p.isdigit()), None)
        glycan  = "_".join(parts[pos_idx + 1:]) if pos_idx is not None else name
        return species, glycan

    folders = [d for d in sorted(os.listdir(OUTDIR))
               if os.path.isdir(os.path.join(OUTDIR, d))
               and os.path.exists(os.path.join(OUTDIR, d, "all.pdb"))]

    all_stats = []
    for folder in folders:
        species, glycan = parse_folder(folder)
        src  = os.path.join(OUTDIR, folder, "all.pdb")
        dst  = os.path.join(OUTDIR, f"{species}_{glycan}.pdb")
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
            print(f"  {folder} -> {species}_{glycan}.pdb  ({os.path.getsize(dst)/1024/1024:.1f} MB)")
        else:
            print(f"  {species}_{glycan}.pdb 已存在，跳过")

        csv = os.path.join(OUTDIR, folder, "ensemble_stats.csv")
        if os.path.exists(csv) and os.path.getsize(csv) > 0:
            df = pd.read_csv(csv)
            df.insert(0, "物种",       species)
            df.insert(1, "Glycan_ID", glycan)
            all_stats.append(df)

    if all_stats:
        merged   = pd.concat(all_stats, ignore_index=True)
        xlsx_out = os.path.join(OUTDIR, "ensemble_stats_all.xlsx")
        with pd.ExcelWriter(xlsx_out, engine="openpyxl") as writer:
            merged.to_excel(writer, index=False, sheet_name="ensemble_stats")
            ws = writer.sheets["ensemble_stats"]
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = \
                    min(max(len(str(c.value or "")) for c in col) + 2, 50)
        print(f"\n✅ 合并统计 ({len(merged)} 行) -> {xlsx_out}")

    # 删除子文件夹
    removed = 0
    for d in folders:
        dp = os.path.join(OUTDIR, d)
        shutil.rmtree(dp, ignore_errors=True)
        removed += 1
    print(f"✅ 已删除 {removed} 个子文件夹")


# ══════════════════════════════════════════════════════════════
#  步骤 3b: NeuAc 异构体建模 (补充)
#  原脚本: reglyco_neuac_isoforms.py
# ══════════════════════════════════════════════════════════════
def step3b_neuac_isoforms():
    """
    补充建模 HexNAc(5)Hex(6)NeuAc(1) 的全部 GlycoShape 异构体。
    蛋白: A0A2I0MWA2 (Columba), 位点 97_A
    GS00060 已在步骤3完成，此处补充其余4个 (GS00061/62/496/917)。
    GS00889 beta=None，GlycoShape 无 MD 构象库，无法建模，跳过。
    """
    print("\n" + "="*60)
    print("步骤 3b: NeuAc 异构体建模 (GS00061/62/496/917)")
    print("="*60)

    PROT_ID  = "A0A2I0MWA2"
    POSITION = 97
    TASKS    = [
        ("GS00061", "G96963QG", "NeuAc-a2-3 on alpha1-3 arm"),
        ("GS00062", "G71405MT", "NeuAc-a2-3 beta1-6-GlcNAc antenna"),
        ("GS00496", "G79578JB", "NeuAc-a2-3 on alpha1-6 arm (tri-ant)"),
        ("GS00917", "G01571OF", "NeuAc-a2-3 Gal-GlcNAc alpha1-6, alpha1-3 undecorated"),
    ]

    os.makedirs(OUTDIR, exist_ok=True)
    uuid, avail = _create_session(PROT_ID)
    avail_map   = {s["residueID"]: s for s in avail}
    if POSITION not in avail_map:
        print(f"❌ 位点 {POSITION} 不在结构中，跳过")
        return

    chain    = avail_map[POSITION]["residueChain"]
    site_key = f"{POSITION}_{chain}"
    print(f"Session: {uuid}  site_key: {site_key}\n")

    summary = []
    for gs_id, beta_ac, desc in TASKS:
        folder_name = f"Columba_A0A2I0MWA2_97_NeuAc1_{gs_id}"
        out_folder  = os.path.join(OUTDIR, folder_name)
        if os.path.exists(os.path.join(out_folder, "all.pdb")):
            print(f"  ⏭️  {folder_name} 已存在，跳过")
            summary.append((gs_id, beta_ac, "已存在"))
            continue
        print(f"\n  任务: {gs_id}  beta={beta_ac}  ({desc})")
        try:
            job_uuid = _submit_job(uuid, site_key, beta_ac)
            job_d    = _poll_job(job_uuid)
            status   = job_d.get("status", "unknown")
            if status == "completed":
                _download_results(job_d, out_folder)
                summary.append((gs_id, beta_ac, "completed"))
            else:
                print(f"  ❌ {status}")
                summary.append((gs_id, beta_ac, f"失败:{status}"))
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            summary.append((gs_id, beta_ac, f"异常:{e}"))

    print("\n=== 汇总 ===")
    for gs_id, ac, st in summary:
        print(f"  {gs_id}  {ac:<12}  {st}")


# ══════════════════════════════════════════════════════════════
#  步骤 3c: 补建 GS00635 (Glc₃Man₉)
#  原脚本: reglyco_gs00635.py
# ══════════════════════════════════════════════════════════════
def step3c_gs00635():
    """
    补建 GS00635 (Glc₃Man₉, beta=G11546AX)。
    蛋白: A0A2I0MWA2 (Columba), 位点 97_A。
    此结构在步骤3的批处理中被遗漏（JSON 中 Hex11 行只取了第一个匹配 GS00350）。
    """
    print("\n" + "="*60)
    print("步骤 3c: 补建 GS00635 (Glc₃Man₉)")
    print("="*60)

    PROT_ID     = "A0A2I0MWA2"
    POSITION    = 97
    beta_ac     = "G11546AX"
    folder_name = "Columba_A0A2I0MWA2_97_Hex11_GS00635"
    out_folder  = os.path.join(OUTDIR, folder_name)

    if os.path.exists(os.path.join(out_folder, "all.pdb")):
        print("⏭️  已存在，跳过")
        return

    uuid, avail = _create_session(PROT_ID)
    site        = next((s for s in avail if s["residueID"] == POSITION), None)
    if not site:
        print(f"❌ 位点 {POSITION} 不在结构中"); return
    site_key = f"{POSITION}_{site['residueChain']}"
    print(f"Session: {uuid}  site_key: {site_key}")

    try:
        job_uuid = _submit_job(uuid, site_key, beta_ac)
        print(f"Job: {job_uuid}")
        job_d    = _poll_job(job_uuid)
        status   = job_d.get("status", "unknown")
        if status == "completed":
            _download_results(job_d, out_folder)
            print("✅ 完成!")
        else:
            print(f"❌ 失败: {status}")
    except Exception as e:
        print(f"❌ 异常: {e}")


# ══════════════════════════════════════════════════════════════
#  工具函数 (原 dry_run_check.py / show_isoforms.py /
#             check_neuac.py / test_ensemble_api.py /
#             find_api_url.py)
# ══════════════════════════════════════════════════════════════

def util_dry_run():
    """
    Dry-run: 显示步骤3将提交的全部任务列表，不实际建模。
    原脚本: dry_run_check.py
    """
    print("\n" + "="*60)
    print("Dry-Run: 步骤3任务预览")
    print("="*60)

    xlsx_in = XLSX_UPDATED if os.path.exists(XLSX_UPDATED) else XLSX_ORIG
    xl = pd.ExcelFile(xlsx_in)
    df = xl.parse("糖链")
    df.columns = ["idx","species","protein_id","position","gtype","composition",
                  "AC","note","GS_ID","status"]
    df["AC"] = df["AC"].astype(str).str.strip()

    with open(JSON_AC, encoding="utf-8") as f:
        ac_data = json.load(f)

    gs_ac_map = {}
    for entry in ac_data:
        r     = entry["row"]
        orig  = entry.get("existing_AC", "").strip()
        vlist = entry.get("verified", [])
        if not vlist:
            continue
        exact = any(v.get("glytoucan_beta") == orig for v in vlist)
        if exact:
            gs_ac_map[r] = (orig, orig)
        else:
            first = vlist[0]
            beta  = first.get("glytoucan_beta") or first.get("glytoucan_archetype")
            if beta:
                gs_ac_map[r] = (beta, orig)

    df_valid          = df[df.index.map(lambda i: i in gs_ac_map)].copy()
    df_valid["position"] = df_valid["position"].astype(int)

    print(f"共 {len(df_valid)} 个有效任务:")
    for i, row in df_valid.iterrows():
        use, orig = gs_ac_map[i]
        mark  = "✅" if use == orig else "⚠️"
        arrow = f" -> {use}" if use != orig else ""
        print(f"  Row{i:2d} {str(row['species']):10} {str(row['protein_id']):15} "
              f"位点{row['position']:4} {mark} {orig}{arrow}")
        print(f"       comp={row['composition']}")

    print("\n原始表格中跳过的行（❌ 无GlycoShape匹配）:")
    for i, row in df.iterrows():
        if i not in gs_ac_map:
            print(f"  Row{i:2d} {str(row['species']):10} AC={str(row['AC']):12} "
                  f"{row['composition']}")


def util_show_isoforms():
    """
    打印 glycan_ac_result.json 中每行的 GlycoShape 异构体及其 IUPAC 结构。
    原脚本: show_isoforms.py
    """
    print("\n" + "="*60)
    print("GlycoShape 异构体列表")
    print("="*60)
    with open(JSON_AC, encoding="utf-8") as f:
        data = json.load(f)
    for entry in data:
        r     = entry["row"]
        comp  = entry.get("composition", "")
        ac    = entry.get("existing_AC", "")
        vlist = entry.get("verified", [])
        if not vlist:
            continue
        exact = any(x.get("glytoucan_beta") == ac for x in vlist)
        tag   = "✅" if exact else "⚠️"
        print(f"Row{r:2d} {tag}  comp={comp}  原AC={ac}")
        for x in vlist:
            gs    = x.get("GS_ID", "?")
            beta  = str(x.get("glytoucan_beta", "None"))
            arch  = str(x.get("glytoucan_archetype", "None"))
            iupac = x.get("iupac", "")
            print(f"  GS={gs:8s}  beta={beta:<12s}  arch={arch:<12s}")
            if iupac:
                print(f"  IUPAC={iupac}")
        print()


def util_check_neuac():
    """
    列出 JSON 中所有含 NeuAc 的行及其 GlycoShape 异构体，并检查输出目录中相关文件夹。
    原脚本: check_neuac.py
    """
    print("\n" + "="*60)
    print("NeuAc 相关条目检查")
    print("="*60)
    with open(JSON_AC, encoding="utf-8") as f:
        data = json.load(f)

    print("=== 含 NeuAc 的所有行及 GlycoShape 异构体 ===")
    for entry in data:
        comp = entry.get("composition", "")
        if "NeuAc" not in comp:
            continue
        r    = entry["row"]
        ac   = entry.get("existing_AC", "")
        vlist = entry.get("verified", [])
        print(f"Row{r:2d}  comp={comp}  原AC={ac}")
        if not vlist:
            print("  -> 无 GlycoShape 匹配（已跳过）")
        for x in vlist:
            gs    = x.get("GS_ID", "?")
            beta  = str(x.get("glytoucan_beta", "None"))
            iupac = x.get("iupac", "")[:110]
            print(f"  {gs}  beta={beta:<12s}  IUPAC={iupac}")
        print()

    print("=== 输出目录中 NeuAc 相关文件夹 ===")
    if os.path.isdir(OUTDIR):
        for d in sorted(os.listdir(OUTDIR)):
            dp      = os.path.join(OUTDIR, d)
            has_pdb = os.path.isdir(dp) and os.path.exists(os.path.join(dp, "all.pdb"))
            if "G49478" in d or "NeuAc" in d or "GS000" in d:
                tag = "✅ pdb" if has_pdb else "❌ 无pdb"
                print(f"  {d}  {tag}")


def util_test_api():
    """
    测试 GlycoShape API 连通性：创建 P01012 (Gallus卵清蛋白) session 并提交单个 ensemble job。
    原脚本: test_ensemble_api.py
    """
    print("\n" + "="*60)
    print("API 连通性测试 (P01012, 位点 293)")
    print("="*60)

    r = requests.post(f"{API}/api/sessions", json={"protID": "P01012"}, timeout=60)
    d = r.json()
    session_uuid = d["session_uuid"]
    print(f"Session UUID: {session_uuid}")

    r2        = requests.get(f"{API}/api/sessions/{session_uuid}", timeout=30)
    available = r2.json().get("glycosylation", {}).get("available", [])
    site293   = [s for s in available if s["residueID"] == 293]
    asn_sites = [s for s in available if s["residueName"] == "ASN"]
    print(f"位点 293: {site293}")
    print(f"所有 ASN 位点: {[(s['residueID'], s['residueChain']) for s in asn_sites]}")

    job_payload = {
        "session_uuid":    session_uuid,
        "jobType":         "ensemble",
        "selectedGlycans": {"293_A": "G80966KZ"},
        "parameters": {
            "ensembleSize": 10,
            "calculateSASA": False,
            "seed": 42,
            "outputFormat": "PDB"
        }
    }
    print("\n提交 job:", json.dumps(job_payload, indent=2, ensure_ascii=False))
    r3 = requests.post(f"{API}/api/jobs", json=job_payload, timeout=60)
    print(f"响应 {r3.status_code}: {r3.text[:500]}")


def util_find_api_url():
    """
    从 GlycoShape 前端 JS bundle 中提取 API 地址和 jobType 等关键参数。
    原脚本: find_api_url.py
    """
    import re
    print("\n" + "="*60)
    print("从 JS bundle 提取 API 配置")
    print("="*60)

    r  = requests.get("https://glycoshape.org/assets/index-DDHOntUZ.js", timeout=60)
    js = r.text

    vw_hits = re.findall(r'VW\s*=\s*["\' ]*(https?://[^"\'` ;]+)["\' ]*', js)
    print(f"VW (apiUrl): {vw_hits[:5]}")

    idx = js.find('"/api/jobs"')
    if idx >= 0:
        print("\n=== /api/jobs 上下文 ===")
        print(js[max(0, idx - 1000):idx + 500])

    job_types = re.findall(r'jobType[:\s=]+["\' ]*([a-zA-Z_]+)["\' ]*', js)
    print(f"\njobType 枚举值: {list(dict.fromkeys(job_types))[:10]}")

    for m in re.finditer(r"reglyco/init", js):
        start = max(0, m.start() - 500)
        end   = min(len(js), m.end() + 500)
        print("\n=== reglyco/init 上下文 ===")
        print(js[start:end])
        break


# ══════════════════════════════════════════════════════════════
#  主入口
# ══════════════════════════════════════════════════════════════
STEPS = {
    # 主流程
    "1":  step1_verify_ac,
    "2":  step2_update_excel,
    "3":  step3_run_ensemble,
    "3b": step3b_neuac_isoforms,
    "3c": step3c_gs00635,
    "4":  step4_model_status,
    "5":  step5_flatten,
    # 别名
    "verify_ac":       step1_verify_ac,
    "update_excel":    step2_update_excel,
    "run_ensemble":    step3_run_ensemble,
    "neuac_isoforms":  step3b_neuac_isoforms,
    "gs00635":         step3c_gs00635,
    "model_status":    step4_model_status,
    "flatten":         step5_flatten,
    # 工具
    "dry_run":         util_dry_run,
    "show_isoforms":   util_show_isoforms,
    "check_neuac":     util_check_neuac,
    "test_api":        util_test_api,
    "find_api_url":    util_find_api_url,
}


"""
run_analysis.py
===============
合并了 Re-Glyco Ensemble 的全部非可视化分析流程，共分为4步。

用法：
  python run_analysis.py            # 依次运行所有步骤
  python run_analysis.py --step 1   # 仅运行 Step 1 (APBS + Ensemble SASA)
  python run_analysis.py --step 2   # 仅运行 Step 2 (糖链几何指标)
  python run_analysis.py --step 3   # 仅运行 Step 3 (糖链遮蔽与热点计算)
  python run_analysis.py --step 4   # 仅运行 Step 4 (ANOVA与Duncan统计)
"""


# 强制 UTF-8 输出
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')
warnings.filterwarnings("ignore")

# ==============================================================================
# 全局预设
# ==============================================================================
BASE_DIR    = Path(r'E:\Data\Desktop\Work On\ReGlyco_Ensemble')
PDB_DIR     = BASE_DIR / 'PDB'
CSV_DIR     = BASE_DIR / 'csv'
CSV_DIR.mkdir(exist_ok=True)
OLD_FOLDER  = Path(r"E:\Data\Desktop\Work On\ReGlyco模型")

# 环境与参数
APBS_BINARY      = r"E:\APBS-3.4.1.Windows\bin\apbs.exe"
ENSEMBLE_STEP    = 5
CA2_PROBE_BARE   = 1.00
CA2_PROBE_HYD    = 4.12
APBS_THRESHOLD   = -5.0
SASA_MIN         = 1.0
INTERFACE_CUTOFF = 15.0
SPECIES_ORDER    = ['Gallus', 'Anas', 'Columba']

GLYCAN_CHAIN = 'B'
GLYCAN_RESNAMES = {
    'FUC','GAL','GLA','GLC','GCS','GCU','NAG','NDG','MAN','BMA',
    'SIA','NGA','XYL','FCA','FCB','MMA','MAL','GNS','BCA',
    'NeuAc','Neu5Ac','SLN','ANE','LFU','GlyF','GlyB','4YS','2MS','0MK',
}

def build_name_mapping():
    """解析 A1-A3, C1-C14, G1，短名 => Path"""
    counters = defaultdict(int)
    mapping = {}
    pdbs = sorted((p for p in PDB_DIR.glob('*.pdb') if not p.stem.endswith('_apo')),
                  key=lambda p: (p.stem.split('_')[0], p.name))
    for p in pdbs:
        prefix = p.stem.split('_')[0][0].upper()
        counters[prefix] += 1
        short = f"{prefix}{counters[prefix]}"
        mapping[short] = p
    return mapping

NAME_MAP = build_name_mapping()


# ==============================================================================
# Step 1. APBS 与 Ensemble SASA (原 ensemble_calcium_analysis.py)
# ==============================================================================
COORD_ATOMS = {
    'ASP': ['OD1', 'OD2'], 'GLU': ['OE1', 'OE2'], 'ASN': ['OD1'],
    'GLN': ['OE1'], 'SER': ['OG'], 'THR': ['OG1'], 'TYR': ['OH'],
}
MAX_SASA = {
    "ALA":113.1,"ARG":240.7,"ASN":158.7,"ASP":151.0,"CYS":140.0,"GLN":189.7,"GLU":183.0,
    "GLY":85.0, "HIS":194.0,"ILE":182.0,"LEU":180.0,"LYS":211.4,"MET":204.0,"PHE":210.0,
    "PRO":143.0,"SER":128.3,"THR":146.2,"TRP":255.0,"TYR":229.0,"VAL":160.0,
}

def run_step_1():
    print(f"\n{'='*70}\n[Step 1] APBS 计算与 Ensemble SASA 提取\n{'='*70}")
    
    def load_gly_module():
        spec = importlib.util.spec_from_file_location("glycan_aware_apbs", OLD_FOLDER / "glycan_aware_apbs.py")
        m = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(m)
        m.GLYCAM_CHARGES['GLC'] = m.GLYCAM_CHARGES['MAN'].copy()
        m.GLYCAM_CHARGES['NDG'] = m.GLYCAM_CHARGES['NAG'].copy()
        return m
    
    m_gly = load_gly_module()

    # 1. 提取 MODEL 1 并切分
    def extract_model1(pdb_path, work_dir):
        work_dir.mkdir(parents=True, exist_ok=True)
        prot_lines, gly_lines = [], []
        in_m1, done = False, False
        with open(pdb_path, encoding='utf-8', errors='replace') as f:
            for line in f:
                if done: break
                if line.startswith('MODEL') and not in_m1: in_m1 = True; continue
                if line.startswith('ENDMDL') and in_m1: done = True; continue
                if not in_m1: continue
                rec = line[:6].strip()
                if rec != 'ATOM': continue
                chain = line[21]
                resname = line[17:20].strip()
                if chain == GLYCAN_CHAIN and resname in GLYCAN_RESNAMES:
                    gly_lines.append('HETATM' + line[6:])
                elif chain == GLYCAN_CHAIN: pass
                else: prot_lines.append(line)
        P, G = work_dir / "protein_m1.pdb", work_dir / "glycan_m1.pdb"
        with open(P, 'w') as f: f.writelines(prot_lines); f.write("END\n")
        with open(G, 'w') as f: f.writelines(gly_lines); f.write("END\n")
        return str(P), str(G)

    # 2. 运行单体 PDB 的 APBS
    apbs_results = {}
    for short, pdb in NAME_MAP.items():
        out_csv = CSV_DIR / f"{short}_APBS_glycanAware.csv"
        print(f"  → {short} APBS...")
        if out_csv.exists():
            apbs_results[short] = pd.read_csv(out_csv)
            continue
            
        w_dir = BASE_DIR / f"_apbs_tmp_{short}"
        P, G = extract_model1(pdb, w_dir)
        prot_pqr, gly_pqr, complex_pqr = str(w_dir/"protein.pqr"), str(w_dir/"glycan.pqr"), str(w_dir/"complex.pqr")
        m_gly.run_pdb2pqr(P, prot_pqr)
        if sum(1 for l in open(G) if l.startswith('HETATM')) > 0:
            m_gly.glycan_to_pqr(G, gly_pqr)
        else:
            open(gly_pqr, 'w').close()
        m_gly.merge_pqr(prot_pqr, gly_pqr, complex_pqr)
        
        apbs_in = str(w_dir / "complex.in")
        dx_rel = f"{short}_pot"
        dx_path = str(w_dir / f"{dx_rel}.dx")
        c, L, ns = m_gly.apbs_extent(complex_pqr)
        m_gly.write_apbs_input(apbs_in, "complex.pqr", dx_rel, c, L, ns)
        
        subprocess.run([APBS_BINARY, apbs_in], capture_output=True, cwd=str(w_dir))
        comb_pdb = str(w_dir / "combined_m1.pdb")
        with open(comb_pdb, 'w') as f:
            f.writelines(l for l in open(P) if l.startswith('ATOM'))
            f.writelines(l for l in open(G) if l.startswith('HETATM'))
            f.write("END\n")
        df = m_gly.interpolate_to_residues(dx_path, comb_pdb, include_glycan=True)
        df.to_csv(out_csv, index=False)
        apbs_results[short] = df

    # 3. 对 50 模型做 Ensemble SASA
    ensemble_sasa = {}
    print(f"\n  [Ensemble SASA - 取样间隔 {ENSEMBLE_STEP}]")
    for short, pdb in NAME_MAP.items():
        out_sasa = CSV_DIR / f"{short}_ensemble_sasa.csv"
        print(f"  → {short} Ensemble SASA...")
        if out_sasa.exists():
            ensemble_sasa[short] = pd.read_csv(out_sasa)
            continue
            
        blocks, cur, in_m = [], [], False
        for l in open(pdb, errors='replace'):
            if l.startswith('MODEL'): in_m=True; cur=[]
            elif l.startswith('ENDMDL') and in_m: blocks.append(cur); cur=[]; in_m=False
            elif in_m: cur.append(l)
        
        sampled = blocks[::ENSEMBLE_STEP]
        sasa_recs = defaultdict(lambda: defaultdict(list))
        import io
        for i, blk in enumerate(sampled):
            prot_l = [l for l in blk if l.startswith('ATOM') and l[21]!=GLYCAN_CHAIN]
            struct = PDBParser(QUIET=True).get_structure('m', io.StringIO(''.join(prot_l)+'END\n'))
            model_obj = next(struct.get_models())
            for pr, lbl in [(1.4, 'water'), (CA2_PROBE_BARE, 'ca2_bare'), (CA2_PROBE_HYD, 'ca2_hyd')]:
                ShrakeRupley(probe_radius=pr, n_points=100).compute(model_obj, level='A')
                for ch in model_obj:
                    for res in ch:
                        if not is_aa(res): continue
                        key = (ch.id, res.resname, res.id[1])
                        c_sasa = sum(a.sasa for a in res if a.name in COORD_ATOMS.get(res.resname,[]) and hasattr(a,'sasa'))
                        sasa_recs[key][f'coord_sasa_{lbl}'].append(c_sasa)
                        if lbl=='water': sasa_recs[key]['res_sasa'].append(sum(a.sasa for a in res if hasattr(a,'sasa')))
                        
        rows = []
        for (c,rn,rs), vals in sasa_recs.items():
            row = {'Chain': c, 'ResName': rn, 'ResSeq': rs}
            for lbl in ['coord_sasa_water', 'coord_sasa_ca2_bare', 'coord_sasa_ca2_hyd', 'res_sasa']:
                arr = np.array(vals[lbl])
                row[lbl+'_mean'] = arr.mean() if len(arr) else np.nan
                row[lbl+'_std']  = arr.std()  if len(arr) else np.nan
            ms = MAX_SASA.get(rn, 150)
            row['RelSASA_mean'] = min(row['res_sasa_mean']/ms, 1.0) if ms>0 else 0
            # 重命名 res_sasa 成与旧代码匹配的兼容格式
            row['res_sasa_water_mean'] = row.pop('res_sasa_mean')
            row.pop('res_sasa_std')
            rows.append(row)
            
        df = pd.DataFrame(rows)
        df.to_csv(out_sasa, index=False)
        ensemble_sasa[short] = df

    print(f"\nStep 1 完成。APBS & SASA 文件已存入 {CSV_DIR}")


# ==============================================================================
# Step 2. 糖链构象与几何 (原 glycan_ensemble_stats.py)
# ==============================================================================
def run_step_2():
    print(f"\n{'='*70}\n[Step 2] 糖链构象与几何指标提取 (基于距离计算)\n{'='*70}")
    
    def parse_ens(p):
        md, cur, pca, gall, gres = {}, None, [], [], defaultdict(list)
        def flush(m): md[m] = {'p':np.array(pca),'g':np.array(gall),'r':dict(gres)}
        for l in open(p, errors='replace'):
            rec = l[:6].strip()
            if rec=='MODEL':
                if cur: flush(cur)
                cur = int(l.split()[1]) if len(l.split())>1 else 1
                pca, gall, gres = [], [], defaultdict(list)
            elif rec=='ENDMDL' and cur: flush(cur); cur=None; pca,gall,gres=[],[],defaultdict(list)
            elif rec in ('ATOM','HETATM') and len(l)>=54:
                ele = l[76:78].strip() if len(l)>77 else ''
                if ele.upper() in ('H','D'): continue
                if not ele and l[12:16].strip().startswith('H'): continue
                try: x,y,z = float(l[30:38]),float(l[38:46]),float(l[46:54])
                except: continue
                is_g = (l[21]==GLYCAN_CHAIN) or (l[17:20].strip() in GLYCAN_RESNAMES and rec=='HETATM')
                if is_g:
                    idx = len(gall); gall.append([x,y,z])
                    gres[(l[21], l[17:20].strip(), l[22:26].strip())].append(idx)
                elif rec=='ATOM' and l[12:16].strip()=='CA':
                    pca.append([x,y,z])
        if cur and (pca or gall): flush(cur)
        return md

    rows = []
    for short, pdb in sorted(NAME_MAP.items()):
        print(f"  → {short} ({pdb.stem})")
        md = parse_ens(pdb)
        for mn, d in md.items():
            g, p, r = d['g'], d['p'], d['r']
            if len(g)==0: continue
            
            rg = np.sqrt(np.mean(np.sum((g - g.mean(axis=0))**2, axis=1)))
            c0, c1 = list(r.keys())[0], list(r.keys())[-1]
            e2e = np.linalg.norm(g[r[c1]].mean(axis=0) - g[r[c0]].mean(axis=0)) if len(r)>1 else np.nan
            dist = np.linalg.norm(g.mean(axis=0) - p.mean(axis=0)) if len(p)>0 else np.nan
            mdist = cdist(g, p).min() if len(p)>0 else np.nan
            
            rows.append({
                'structure': pdb.stem, 'species': pdb.stem.split('_')[0], 'model': mn,
                'glycan_atoms': len(g), 'glycan_rg': rg, 'glycan_dist': dist,
                'glycan_end2end': e2e, 'glycan_min_dist_to_ca': mdist
            })
            
    df = pd.DataFrame(rows)
    out1 = CSV_DIR / 'glycan_conformation_detail.csv'
    df.to_csv(out1, index=False)
    print(f"\nStep 2 完成。已保存 {out1} (共 {len(df)} 行)")


# ==============================================================================
# Step 3. 糖链遮蔽效应与 Net Hotspot 计算 (原 glycan_shielding_hotspot.py)
# ==============================================================================
def run_step_3():
    print(f"\n{'='*70}\n[Step 3] 全结构 SASA 糖链热点遮蔽计算 (n_points=20)\n{'='*70}")
    
    out_csv = CSV_DIR / 'hotspot_per_conformation.csv'
    done_set = set()
    exists_rows = []
    if out_csv.exists():
        old = pd.read_csv(out_csv)
        done_set = set(old['short_name'].unique())
        exists_rows.append(old)
        print(f"  [*] 发现断点数据，已完成结构: {done_set}...")

    import io
    parser = PDBParser(QUIET=True)
    all_rows = list(exists_rows)

    for short, pdb in sorted(NAME_MAP.items()):
        if short in done_set: continue
        print(f"  → {short} 构象集分析中...")
        
        apbs_csv = CSV_DIR / f"{short}_APBS_glycanAware.csv"
        if not apbs_csv.exists():
            print(f"   [!] 找不到 APBS 基础文件 {apbs_csv.name}，跳过。")
            continue
            
        apbs_df = pd.read_csv(apbs_csv)
        apbs_df['ResSeq'] = apbs_df['ResSeq'].astype(str).str.strip()
        apbs_dict = dict(zip(apbs_df['ResSeq'], apbs_df['APBS_kT_e']))
        sasa_alone = dict(zip(apbs_df['ResSeq'], apbs_df['SASA_A2']))
        cand_seqs = set(apbs_df.loc[apbs_df['APBS_kT_e'] < APBS_THRESHOLD, 'ResSeq'])
        n_cand = len(cand_seqs)

        blocks, cur, in_m = [], [], False
        for l in open(pdb, errors='replace'):
            if l.startswith('MODEL'): in_m=True; cur=[]
            elif l.startswith('ENDMDL') and in_m: blocks.append(cur); cur=[]; in_m=False
            elif in_m: cur.append(l)
            
        # 界面的残基判断
        g_coords = [[float(l[30:38]),float(l[38:46]),float(l[46:54])] for l in blocks[0] 
                    if l[:6].strip() in ('ATOM','HETATM') and len(l)>54 and 
                    (l[21]==GLYCAN_CHAIN or (l[17:20].strip() in GLYCAN_RESNAMES and l.startswith('HETATM')))]
        g_arr = np.array(g_coords)
        iface = set()
        if len(g_arr) > 0:
            for _, row in apbs_df.iterrows():
                ca = np.array([row['CA_x'], row['CA_y'], row['CA_z']])
                if np.min(np.linalg.norm(g_arr - ca, axis=1)) <= INTERFACE_CUTOFF:
                    iface.add(str(row['ResSeq']))
                    
        curr_struct_rows = []
        t0 = time.time()
        for idx, blk in enumerate(blocks):
            pst = 'CRYST1    1.000    1.000    1.000  90.00  90.00  90.00 P 1           1\n' + ''.join(blk) + 'END\n'
            mobj = next(parser.get_structure('m', io.StringIO(pst)).get_models())
            ShrakeRupley(probe_radius=1.4, n_points=20).compute(mobj, level='R')
            
            f_sasa = {}
            for ch in mobj:
                if ch.id == 'A':
                    for res in ch:
                        if is_aa(res): f_sasa[str(res.id[1])] = res.sasa
                        
            n_hot, n_shield, sh_delta, h_sasa = 0, 0, [], []
            for rs in cand_seqs:
                fv = f_sasa.get(rs, sasa_alone.get(rs, 0))
                av = sasa_alone.get(rs, 0)
                if fv >= SASA_MIN:
                    n_hot += 1; h_sasa.append(fv)
                
                dt = av - fv
                if av >= SASA_MIN and dt > 5.0:
                    n_shield += 1; sh_delta.append(dt)
                    
            iface_f = [f_sasa.get(s, np.nan) for s in iface]
            iface_a = [sasa_alone.get(s, np.nan) for s in iface]
            
            curr_struct_rows.append({
                'short_name': short, 'structure': pdb.stem, 'species': pdb.stem.split('_')[0],
                'model': idx+1, 'n_candidates': n_cand, 'n_hotspots': n_hot,
                'hotspot_frac': n_hot/n_cand if n_cand else np.nan, 'n_shielded_cands': n_shield,
                'shielded_sasa_delta': np.mean(sh_delta) if sh_delta else 0,
                'hotspot_sasa_mean': np.mean(h_sasa) if h_sasa else np.nan,
                'iface_full_sasa': np.nanmean(iface_f) if len(iface_f) else np.nan,
                'iface_alone_sasa': np.nanmean(iface_a) if len(iface_a) else np.nan,
                'iface_shielding': (np.nanmean(iface_a)-np.nanmean(iface_f)) if len(iface_a) else np.nan
            })
            
        print(f"      [{time.time()-t0:.1f}s] {len(blocks)} consts; "
              f"avg net hotspots = {np.mean([r['n_hotspots']-r['n_shielded_cands'] for r in curr_struct_rows]):.1f}")
              
        df_sub = pd.DataFrame(curr_struct_rows)
        all_rows.append(df_sub)
        pd.concat(all_rows, ignore_index=True).to_csv(out_csv, index=False)

    df_full = pd.concat(all_rows, ignore_index=True)
    sum_rows = []
    for (sh, st, sp), g in df_full.groupby(['short_name','structure','species']):
        sum_rows.append({
            'short_name': sh, 'species': sp, 'n_models': len(g),
            'hotspot_mean': g['n_hotspots'].mean(),
            'shielded_cand_mean': g['n_shielded_cands'].mean(),
            'iface_shielding_mean': g['iface_shielding'].mean()
        })
    pd.DataFrame(sum_rows).to_csv(CSV_DIR / 'hotspot_summary.csv', index=False)
    print(f"\nStep 3 完成。热点分布数据保存在 {out_csv}")


# ==============================================================================
# Step 4. Duncan's Multi-Range Test (原 _stats_report.py)
# ==============================================================================
def duncan_mrt(groups, labels, alpha=0.05):
    from scipy.stats import studentized_range
    ns = np.array([len(g) for g in groups], dtype=float)
    k, df_err = len(groups), int(ns.sum() - len(groups))
    MSE = sum(np.sum((g - g.mean())**2) for g in groups) / df_err
    n_h = k / np.sum(1.0 / ns)
    s_y = np.sqrt(MSE / n_h)
    
    means = np.array([g.mean() for g in groups])
    order = np.argsort(means)[::-1]
    s_means, s_labels = means[order], [labels[i] for i in order]
    
    res = {}
    for i in range(k):
        for j in range(i+1, k):
            p = j - i + 1
            LSR = studentized_range.ppf(1-alpha, p, df_err) * s_y
            diff = s_means[i] - s_means[j]
            res[(s_labels[i], s_labels[j])] = {'diff':diff, 'LSR':LSR, 'sig':diff>=LSR}
    return res

def run_step_4():
    print(f"\n{'='*70}\n[Step 4] 统计学评估 (One-Way ANOVA & Duncan Multiple Range Test)\n{'='*70}")
    
    h_file = CSV_DIR / 'hotspot_per_conformation.csv'
    g_file = CSV_DIR / 'glycan_conformation_detail.csv'
    if not h_file.exists() or not g_file.exists():
        print("缺少统计必须的 CSV 文件，请先运行 Step 2 和 Step 3。")
        return
        
    df_h = pd.read_csv(h_file)
    df_g = pd.read_csv(g_file)
    df_h['net_accessible'] = df_h['n_hotspots'] - df_h['n_shielded_cands']
    
    cmp_list = [
        (df_h, 'net_accessible', 'Net Accessible Hotspots (Phenotypic Indicator)'),
        (df_h, 'n_hotspots', 'Total Hotspots (Exposed SASA > 1 Å²)'),
        (df_h, 'n_shielded_cands', 'Glycan-Shielded Hotspots'),
        (df_h, 'iface_shielding', 'Interface SASA Shielding (Å²)'),
        (df_g, 'glycan_rg', 'Glycan Radius of Gyration (Å)'),
        (df_g, 'glycan_end2end', 'Glycan End-to-End Distance (Å)')
    ]
    
    for df, metric, name in cmp_list:
        print(f"\n➤ {name}")
        groups = [df.loc[df.species == s, metric].dropna().values for s in SPECIES_ORDER]
        
        # ANOVA
        from scipy.stats import f_oneway
        F, pv = f_oneway(*groups)
        sig = '***' if pv<0.001 else '**' if pv<0.01 else '*' if pv<0.05 else 'ns'
        print(f"  ANOVA F={F:.2f}, p={pv:.2e} {sig}")
        
        for s, g in zip(SPECIES_ORDER, groups):
            print(f"    {s:10s}: {g.mean():6.2f} ± {g.std():5.2f}")
            
        # Duncan
        if pv < 0.05:
            dres = duncan_mrt(groups, SPECIES_ORDER)
            print("  [Duncan DMRT, alpha=0.05]")
            for pair, d in dres.items():
                print(f"    {pair[0]} vs {pair[1]}: diff={d['diff']:.2f}, LSR={d['LSR']:.2f} 显著={d['sig']}")
                
    print("\nStep 4 完成。")


# ==============================================================================
# 入口组装
# ==============================================================================

def pipeline_main():
    parser = argparse.ArgumentParser(
        description="糖蛋白 Re-Glyco Ensemble 完整流程",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
主流程步骤:
  1 / verify_ac      — 从 GlycoShape 验证/查找糖链 AC
  2 / update_excel   — 将 AC 验证结果写入 Excel
  3 / run_ensemble   — 批量 Re-Glyco Ensemble 建模
  3b / neuac_isoforms — NeuAc 异构体补充建模
  3c / gs00635       — 补建 GS00635 (Glc₃Man₉)
  4 / model_status   — 更新 Excel 建模状态列
  5 / flatten        — 整理输出（重命名 PDB、合并统计）
  all                — 依次执行 1→2→3→3b→3c→4→5

工具:
  dry_run        — 预览步骤3任务列表
  show_isoforms  — 打印 JSON 中所有异构体
  check_neuac    — 检查 NeuAc 相关条目
  test_api       — API 连通性测试
  find_api_url   — 从 JS bundle 提取 API 地址
        """
    )
    parser.add_argument("step", nargs="?", default="all",
                        choices=list(STEPS.keys()) + ["all"],
                        help="要运行的步骤 (默认: all)")
    args = parser.parse_args()

    if args.step == "all":
        for fn in [step1_verify_ac, step2_update_excel, step3_run_ensemble,
                   step3b_neuac_isoforms, step3c_gs00635,
                   step4_model_status, step5_flatten]:
            fn()
    else:
        STEPS[args.step]()

    print("\n全部完成。")

def analyze_main():
    parser = argparse.ArgumentParser(description="ReGlyco Pipeline")
    parser.add_argument('--step', type=str, default='all', choices=['1', '2', '3', '4', 'all'])
    args = parser.parse_args()

    # 此处不需要再显示长段 log/txt，这是前次 run 的残存文件，
    # 若存在，顺便予以说明（也可让用户自行静默删除或仅在此提示）
    
    s = args.step
    if s in ['1', 'all']: run_step_1()
    if s in ['2', 'all']: run_step_2()
    if s in ['3', 'all']: run_step_3()
    if s in ['4', 'all']: run_step_4()
    
    if s == 'all':
        print("\n\n🎉 1-4 核心分析已全部完成！")
        print("相关图表绘制可继续通过以下文件单独生成：")
        print("  Fig_ensemble_visualization.py")
        print("  Fig_glycan_ensemble_stats_visualization.py\n")
if __name__ == '__main__':
    if len(sys.argv)>1 and sys.argv[1]=='pipeline':
        sys.argv.pop(1)
        pipeline_main()
    elif len(sys.argv)>1 and sys.argv[1]=='analyze':
        sys.argv.pop(1)
        analyze_main()
    else:
        print("Usage: python ReGlyco_computation_all.py [pipeline or analyze] ...")
