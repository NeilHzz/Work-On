"""
export_data_tables.py
=====================
将四张图的所有计算数据导出为 Excel 文件（每张图一个 .xlsx，每个子图一个 Sheet）

输出文件：
  Figure1_Glycan_Conformational_Diversity.xlsx
  Figure2_Hotspot_Count_Trajectory.xlsx
  Figure3_Hotspot_Accessibility.xlsx
  Figure4_APBS_Calcium_Ensemble.xlsx
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import numpy as np
import pandas as pd
from scipy import stats
from pathlib import Path
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

FOLDER  = Path(r"E:\Data\Desktop\Work On\ReGlyco_Ensemble")
CSV_DIR = FOLDER / "csv"

SPECIES_ORDER = ['Gallus', 'Anas', 'Columba']
CARBOXYL      = ['ASP', 'GLU']

GLYC_ORDER = ['G1', 'A1', 'A2', 'A3',
              'C1',  'C2',  'C3',  'C4',  'C5',  'C6',  'C7',
              'C8',  'C9',  'C10', 'C11', 'C12', 'C13', 'C14']
APO_ORDER  = ['G1_apo', 'A1_apo', 'C1_apo']

# 物种颜色（ARGB 十六进制）
SP_HEADER_COLOR = {
    'Gallus':  'FFB54664',
    'Anas':    'FF7895C1',
    'Columba': 'FFF0C284',
}
HEADER_GRAY  = 'FF4A4A4A'
SUBHDR_GRAY  = 'FFD9D9D9'
WHITE        = 'FFFFFFFF'

# ── 样式工具 ─────────────────────────────────────────────────────────────────

def _border(thin=True):
    s = Side(style='thin' if thin else 'medium', color='FFB0B0B0')
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=HEADER_GRAY, font_color=WHITE, bold=True, size=10,
                 value=None):
    if value is not None:
        cell.value = value
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=bold, color=font_color, size=size, name='Calibri')
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    cell.border = _border()

def style_subheader(cell, bg=SUBHDR_GRAY):
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(bold=True, size=9, name='Calibri')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _border()

def style_data(cell):
    cell.font = Font(size=9, name='Calibri')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _border(thin=True)

def style_sig(cell, p):
    """根据 p 值着色：***=深红, **=橙, *=黄, ns=浅灰"""
    colors = {'***': 'FFFF4444', '**': 'FFFF9933',
              '*':   'FFFFFF99', 'ns': 'FFF5F5F5'}
    label  = ('***' if p < 0.001 else '**' if p < 0.01
               else '*' if p < 0.05 else 'ns')
    cell.value = f"{label}\n(p={p:.4f})"
    cell.fill  = PatternFill('solid', fgColor=colors[label])
    cell.font  = Font(size=9, name='Calibri', bold=(label != 'ns'))
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    cell.border = _border()

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_df_to_sheet(ws, df, start_row=1, start_col=1,
                      header_bg=HEADER_GRAY, freeze=True):
    """将 DataFrame 写入工作表，带样式标题行"""
    cols = list(df.columns)
    for j, c in enumerate(cols, start_col):
        cell = ws.cell(row=start_row, column=j, value=str(c))
        style_header(cell, bg=header_bg)
        set_col_width(ws, j, max(len(str(c)) + 4, 14))
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, start_col):
            cell = ws.cell(row=i, column=j, value=val)
            style_data(cell)
    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    return start_row + len(df) + 1

def add_stats_block(ws, data_dict, label_row, start_col=1):
    """为一个 {group: array} 字典写一个均值/SD/n/CV% 统计块"""
    headers = ['Group', 'n', 'Mean', 'Std', 'Median', 'Q25', 'Q75',
               'Min', 'Max', 'CV%']
    for j, h in enumerate(headers, start_col):
        style_header(ws.cell(row=label_row, column=j, value=h))
        set_col_width(ws, j, 12)
    row = label_row + 1
    for grp, arr in data_dict.items():
        arr = np.asarray(arr, dtype=float)
        vals = [grp, len(arr),
                round(float(np.mean(arr)), 4),
                round(float(np.std(arr, ddof=1)), 4) if len(arr) > 1 else 0,
                round(float(np.median(arr)), 4),
                round(float(np.percentile(arr, 25)), 4),
                round(float(np.percentile(arr, 75)), 4),
                round(float(np.min(arr)), 4),
                round(float(np.max(arr)), 4),
                round(float(np.std(arr, ddof=1) / np.mean(arr) * 100), 2)
                    if np.mean(arr) != 0 and len(arr) > 1 else 0]
        for j, v in enumerate(vals, start_col):
            sp_match = grp if grp in SP_HEADER_COLOR else None
            cell = ws.cell(row=row, column=j, value=v)
            if sp_match and j == start_col:
                cell.fill = PatternFill('solid', fgColor=SP_HEADER_COLOR[sp_match])
                cell.font = Font(bold=True, size=9, color=WHITE, name='Calibri')
            else:
                style_data(cell)
        row += 1
    return row + 1

def add_mannwhitney_block(ws, data_dict, label_row, start_col=1, title=''):
    """写物种两两 Mann-Whitney U 检验结果"""
    ws.cell(row=label_row, column=start_col,
            value=f'Pairwise Mann–Whitney U test — {title}')
    style_subheader(ws.cell(row=label_row, column=start_col))
    ws.merge_cells(start_row=label_row, start_column=start_col,
                   end_row=label_row, end_column=start_col + 3)
    label_row += 1
    for j, h in enumerate(['Group A', 'Group B', 'U statistic', 'p-value / label'],
                           start_col):
        style_header(ws.cell(row=label_row, column=j, value=h))
        set_col_width(ws, j, 16)
    row = label_row + 1
    groups = list(data_dict.keys())
    for i in range(len(groups)):
        for k in range(i+1, len(groups)):
            a, b = groups[i], groups[k]
            ga, gb = np.asarray(data_dict[a]), np.asarray(data_dict[b])
            u, p = stats.mannwhitneyu(ga, gb, alternative='two-sided')
            ws.cell(row=row, column=start_col).value   = a
            ws.cell(row=row, column=start_col + 1).value = b
            ws.cell(row=row, column=start_col + 2).value = round(float(u), 2)
            style_sig(ws.cell(row=row, column=start_col + 3), p)
            for j in range(start_col, start_col + 3):
                style_data(ws.cell(row=row, column=j))
            row += 1
    return row + 1


# ════════════════════════════════════════════════════════════════════════════════
# FIGURE 1 — Glycan Conformational Diversity
# ════════════════════════════════════════════════════════════════════════════════

def export_figure1():
    out = FOLDER / "Figure1_Glycan_Conformational_Diversity.xlsx"
    detail  = pd.read_csv(CSV_DIR / 'glycan_conformation_detail.csv')
    summary = pd.read_csv(CSV_DIR / 'glycan_species_summary.csv')

    with pd.ExcelWriter(out, engine='openpyxl') as writer:

        # ── Sheet 1: Raw conformational detail ──────────────────────────────
        detail_out = detail[['species','structure','model',
                              'glycan_rg','glycan_end2end',
                              'glycan_dist','glycan_min_dist_to_ca']].copy()
        detail_out.columns = ['Species','Structure','Model',
                               'Glycan Rg (Å)','End-to-End Distance (Å)',
                               'Glycan–Protein Distance (Å)',
                               'Min. Distance to Cα (Å)']
        detail_out.to_excel(writer, sheet_name='Raw_Conformation_Data',
                            index=False)
        ws = writer.sheets['Raw_Conformation_Data']
        ws.freeze_panes = 'A2'
        for j, sp in enumerate(detail_out['Species']):
            bg = SP_HEADER_COLOR.get(sp, 'FFFFFFFF')
            ws.cell(row=j+2, column=1).fill = PatternFill('solid', fgColor=bg)
        for j in range(1, len(detail_out.columns)+1):
            style_header(ws.cell(row=1, column=j),
                         value=detail_out.columns[j-1])
            set_col_width(ws, j, 20)

        # ── Sheet 2: Per-structure summary ──────────────────────────────────
        summary_out = summary.rename(columns={
            'structure': 'Structure', 'species': 'Species', 'n': 'N conformations',
            'glycan_rg_mean': 'Rg Mean (Å)', 'glycan_rg_std': 'Rg Std',
            'glycan_rg_cv%': 'Rg CV%',
            'glycan_dist_mean': 'Centroid–Protein Mean (Å)',
            'glycan_dist_std':  'Centroid–Protein Std',
            'glycan_dist_cv%':  'Centroid–Protein CV%',
            'glycan_end2end_mean': 'End-to-End Mean (Å)',
            'glycan_end2end_std':  'End-to-End Std',
            'glycan_end2end_cv%':  'End-to-End CV%',
            'glycan_min_dist_to_ca_mean': 'Min Cα Dist Mean (Å)',
            'glycan_min_dist_to_ca_std':  'Min Cα Dist Std',
            'glycan_min_dist_to_ca_cv%':  'Min Cα Dist CV%',
        })
        summary_out.to_excel(writer, sheet_name='Per-Structure_Summary',
                             index=False)
        ws2 = writer.sheets['Per-Structure_Summary']
        ws2.freeze_panes = 'A2'
        for j in range(1, len(summary_out.columns)+1):
            style_header(ws2.cell(row=1, column=j),
                         value=summary_out.columns[j-1])
            set_col_width(ws2, j, 22)

        # ── Sheet 3: Species-level statistics (per panel) ───────────────────
        metrics = {
            'Glycan Rg (Å)':               'glycan_rg',
            'End-to-End Distance (Å)':     'glycan_end2end',
            'Glycan–Protein Distance (Å)': 'glycan_dist',
            'Min. Distance to Cα (Å)':     'glycan_min_dist_to_ca',
        }
        wb  = writer.book
        ws3 = wb.create_sheet('Species_Statistics')
        row = 1
        for panel_label, (metric_name, col) in zip('ABCD', metrics.items()):
            ws3.cell(row=row, column=1,
                     value=f'Panel {panel_label}: {metric_name}')
            style_subheader(ws3.cell(row=row, column=1))
            ws3.merge_cells(start_row=row, start_column=1,
                            end_row=row, end_column=10)
            row += 1
            data_d = {sp: detail.loc[detail.species==sp, col].dropna().values
                      for sp in SPECIES_ORDER}
            row = add_stats_block(ws3, data_d, row)
            row = add_mannwhitney_block(ws3, data_d, row,
                                       title=metric_name)
            row += 1
        ws3.freeze_panes = 'A1'
        for j in range(1, 11):
            set_col_width(ws3, j, 16)

    print(f"  ✓ {out.name}")


# ════════════════════════════════════════════════════════════════════════════════
# FIGURE 2 — Hotspot Count & Trajectory
# ════════════════════════════════════════════════════════════════════════════════

def export_figure2():
    out = FOLDER / "Figure2_Hotspot_Count_Trajectory.xlsx"
    df  = pd.read_csv(CSV_DIR / 'hotspot_per_conformation.csv')
    hs  = pd.read_csv(CSV_DIR / 'hotspot_summary.csv')
    df['net_accessible'] = df['n_hotspots'] - df['n_shielded_cands']

    with pd.ExcelWriter(out, engine='openpyxl') as writer:

        # ── Sheet A: Total hotspot count (raw) ──────────────────────────────
        panelA = df[['species','short_name','structure','model',
                     'n_candidates','n_hotspots']].rename(columns={
            'species':'Species','short_name':'Short Name',
            'structure':'Structure','model':'Model',
            'n_candidates':'N Candidates','n_hotspots':'Total Hotspot Count'})
        panelA.to_excel(writer, sheet_name='PanelA_Total_Hotspot_Raw', index=False)
        ws = writer.sheets['PanelA_Total_Hotspot_Raw']
        ws.freeze_panes = 'A2'
        for j in range(1, len(panelA.columns)+1):
            style_header(ws.cell(row=1, column=j),
                         value=panelA.columns[j-1])
            set_col_width(ws, j, 20)

        # ── Sheet A-stats ────────────────────────────────────────────────────
        wb  = writer.book
        wsA = wb.create_sheet('PanelA_Statistics')
        row = 1
        wsA.cell(row=row, column=1,
                 value='Panel A: Total Ca²⁺ Hotspot Count (SASA > 1 Å²)')
        style_subheader(wsA.cell(row=row, column=1))
        wsA.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=10)
        row += 1
        data_a = {sp: df.loc[df.species==sp,'n_hotspots'].values
                  for sp in SPECIES_ORDER}
        row = add_stats_block(wsA, data_a, row)
        row = add_mannwhitney_block(wsA, data_a, row,
                                   title='Total Hotspot Count')
        for j in range(1, 11): set_col_width(wsA, j, 16)

        # ── Sheet B: Glycan-shielded hotspot count ───────────────────────────
        panelB = df[['species','short_name','structure','model',
                     'n_shielded_cands','shielded_sasa_delta']].rename(columns={
            'species':'Species','short_name':'Short Name',
            'structure':'Structure','model':'Model',
            'n_shielded_cands':'Glycan-Shielded Hotspot Count',
            'shielded_sasa_delta':'Shielded ΔSASA (Å²) per residue'})
        panelB.to_excel(writer, sheet_name='PanelB_Shielded_Hotspot_Raw',
                        index=False)
        ws = writer.sheets['PanelB_Shielded_Hotspot_Raw']
        ws.freeze_panes = 'A2'
        for j in range(1, len(panelB.columns)+1):
            style_header(ws.cell(row=1, column=j),
                         value=panelB.columns[j-1])
            set_col_width(ws, j, 24)

        wsB = wb.create_sheet('PanelB_Statistics')
        row = 1
        wsB.cell(row=row, column=1,
                 value='Panel B: Glycan-Shielded Hotspot Count (ΔSASA > 5 Å²)')
        style_subheader(wsB.cell(row=row, column=1))
        wsB.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=10)
        row += 1
        data_b = {sp: df.loc[df.species==sp,'n_shielded_cands'].values
                  for sp in SPECIES_ORDER}
        row = add_stats_block(wsB, data_b, row)
        row = add_mannwhitney_block(wsB, data_b, row,
                                   title='Glycan-Shielded Hotspot Count')
        for j in range(1, 11): set_col_width(wsB, j, 16)

        # ── Sheet C: Per-structure trajectory summary ────────────────────────
        hs_out = hs.rename(columns={
            'short_name':'Short Name','structure':'Structure','species':'Species',
            'n_models':'N Models','n_candidates':'N Candidates',
            'hotspot_mean':'Hotspot Count Mean','hotspot_std':'Hotspot Count Std',
            'hotspot_cv%':'Hotspot Count CV%',
            'shielded_cand_mean':'Shielded Count Mean',
            'shielded_cand_std':'Shielded Count Std',
            'iface_shielding_mean':'Interface Shielding Mean (Å²)',
            'iface_shielding_std':'Interface Shielding Std (Å²)',
        })
        hs_out.to_excel(writer, sheet_name='PanelC_Trajectory_Summary',
                        index=False)
        ws = writer.sheets['PanelC_Trajectory_Summary']
        ws.freeze_panes = 'A2'
        for j in range(1, len(hs_out.columns)+1):
            style_header(ws.cell(row=1, column=j),
                         value=hs_out.columns[j-1])
            set_col_width(ws, j, 26)

    print(f"  ✓ {out.name}")


# ════════════════════════════════════════════════════════════════════════════════
# FIGURE 3 — Hotspot Accessibility (5 panels)
# ════════════════════════════════════════════════════════════════════════════════

def export_figure3():
    out = FOLDER / "Figure3_Hotspot_Accessibility.xlsx"
    df  = pd.read_csv(CSV_DIR / 'hotspot_per_conformation.csv')
    df['net_accessible'] = df['n_hotspots'] - df['n_shielded_cands']

    panels_violin = [
        ('PanelA', 'iface_shielding',   'Interface Shielding by Glycan (Å²)'),
        ('PanelB', 'hotspot_sasa_mean', 'Hotspot Residue Mean SASA (Å²)'),
        ('PanelC', 'hotspot_frac',      'Hotspot Fraction (hotspots / candidates)'),
        ('PanelD', 'net_accessible',    'Net Accessible Ca²⁺ Hotspots'),
    ]

    with pd.ExcelWriter(out, engine='openpyxl') as writer:

        # ── Sheets A-D: violin panel raw + stats ────────────────────────────
        for sheet_pfx, col, label in panels_violin:
            raw = df[['species','short_name','structure','model', col]].rename(
                columns={'species':'Species','short_name':'Short Name',
                         'structure':'Structure','model':'Model', col: label})
            raw.to_excel(writer, sheet_name=f'{sheet_pfx}_Raw', index=False)
            ws = writer.sheets[f'{sheet_pfx}_Raw']
            ws.freeze_panes = 'A2'
            for j in range(1, len(raw.columns)+1):
                style_header(ws.cell(row=1, column=j),
                             value=raw.columns[j-1])
                set_col_width(ws, j, 22)

            wb  = writer.book
            wss = wb.create_sheet(f'{sheet_pfx}_Statistics')
            row = 1
            wss.cell(row=row, column=1, value=label)
            style_subheader(wss.cell(row=row, column=1))
            wss.merge_cells(start_row=row, start_column=1,
                            end_row=row, end_column=10)
            row += 1
            data_d = {sp: df.loc[df.species==sp, col].dropna().values
                      for sp in SPECIES_ORDER}
            row = add_stats_block(wss, data_d, row)
            row = add_mannwhitney_block(wss, data_d, row, title=label)
            for j in range(1, 11): set_col_width(wss, j, 16)

        # ── Sheet E: Stacked bar — net accessible vs shielded count ─────────
        rows_e = []
        for sp in SPECIES_ORDER:
            g = df[df.species == sp]
            net = g['net_accessible']
            sh  = g['n_shielded_cands']
            rows_e.append({
                'Species': sp,
                'N conformations': len(g),
                'Net Accessible Count Mean': round(net.mean(), 3),
                'Net Accessible Count Std':  round(net.std(), 3),
                'Net Accessible Count 95%CI': round(net.std()/np.sqrt(len(g))*1.96, 3),
                'Glycan-Shielded Count Mean': round(sh.mean(), 3),
                'Glycan-Shielded Count Std':  round(sh.std(), 3),
                'Total (Net + Shielded) Mean': round((net+sh).mean(), 3),
                'Shielded Fraction (%)': round(sh.mean()/(net+sh).mean()*100, 2),
            })
        dfe = pd.DataFrame(rows_e)
        dfe.to_excel(writer, sheet_name='PanelE_Stacked_Count', index=False)
        wse = writer.sheets['PanelE_Stacked_Count']
        wse.freeze_panes = 'A2'
        for j in range(1, len(dfe.columns)+1):
            style_header(wse.cell(row=1, column=j),
                         value=dfe.columns[j-1])
            set_col_width(wse, j, 26)
        # p-value block
        row = len(dfe) + 4
        wse.cell(row=row, column=1,
                 value='Pairwise Mann–Whitney U: Net Accessible Count')
        style_subheader(wse.cell(row=row, column=1))
        wse.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=4)
        data_e = {sp: df.loc[df.species==sp,'net_accessible'].values
                  for sp in SPECIES_ORDER}
        add_mannwhitney_block(wse, data_e, row+1, title='Net Accessible Count')

        # ── Sheet F: Stacked bar — SASA ──────────────────────────────────────
        rows_f = []
        for sp in SPECIES_ORDER:
            g = df[df.species == sp]
            net = g['iface_full_sasa']
            sh  = g['iface_shielding']
            rows_f.append({
                'Species': sp,
                'N conformations': len(g),
                'Net Accessible SASA Mean (Å²)': round(net.mean(), 3),
                'Net Accessible SASA Std':        round(net.std(), 3),
                'Net Accessible SASA 95%CI':      round(net.std()/np.sqrt(len(g))*1.96, 3),
                'Glycan-Shielded SASA Mean (Å²)': round(sh.mean(), 3),
                'Glycan-Shielded SASA Std':        round(sh.std(), 3),
                'Total SASA Mean (Å²)':            round((net+sh).mean(), 3),
                'Shielded SASA Fraction (%)':      round(sh.mean()/(net+sh).mean()*100, 2),
            })
        dff = pd.DataFrame(rows_f)
        dff.to_excel(writer, sheet_name='PanelF_Stacked_SASA', index=False)
        wsf = writer.sheets['PanelF_Stacked_SASA']
        wsf.freeze_panes = 'A2'
        for j in range(1, len(dff.columns)+1):
            style_header(wsf.cell(row=1, column=j),
                         value=dff.columns[j-1])
            set_col_width(wsf, j, 28)
        row = len(dff) + 4
        wsf.cell(row=row, column=1,
                 value='Pairwise Mann–Whitney U: Net Accessible SASA')
        style_subheader(wsf.cell(row=row, column=1))
        wsf.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=4)
        data_f = {sp: df.loc[df.species==sp,'iface_full_sasa'].values
                  for sp in SPECIES_ORDER}
        add_mannwhitney_block(wsf, data_f, row+1, title='Net Accessible SASA')

    print(f"  ✓ {out.name}")


# ════════════════════════════════════════════════════════════════════════════════
# FIGURE 4 — APBS Ensemble Calcium Analysis
# ════════════════════════════════════════════════════════════════════════════════

def export_figure4():
    out = FOLDER / "Figure4_APBS_Calcium_Ensemble.xlsx"

    summary = pd.read_csv(FOLDER / 'summary_ensemble.csv', encoding='utf-8-sig')
    summary['State'] = summary['IsApo'].map({True: 'Deglycosylated (Apo)',
                                             False: 'Glycosylated'})

    # 读每个结构的 carboxylate SASA
    def read_carboxyl_sasa(names, suffix):
        rows = []
        for n in names:
            base  = n.replace('_apo','')
            fname = f"{base}_{suffix}" if '_apo' not in n else f"{base}_apo_APBS.csv"
            fpath = CSV_DIR / fname
            if fpath.exists():
                df = pd.read_csv(fpath)
                sel = df[(df['Type']=='Protein') &
                         (df['ResName'].isin(CARBOXYL)) &
                         (df['SurfaceLabel']=='Surface')]
                rows.append({'Name': n,
                             'Species': 'Gallus' if n.startswith('G') else
                                        ('Anas' if n.startswith('A') else 'Columba'),
                             'State': 'Deglycosylated (Apo)' if '_apo' in n
                                      else 'Glycosylated',
                             'Carboxylate Surface SASA (Å²)': sel['SASA_A2'].sum()})
        return rows

    sasa_rows = read_carboxyl_sasa(GLYC_ORDER, 'APBS_glycanAware.csv')
    sasa_rows += read_carboxyl_sasa(APO_ORDER,  'APBS_glycanAware.csv')
    sasa_df = pd.DataFrame(sasa_rows)

    # 读取所有 APBS 表面值（用于 Panel D violin）
    apbs_rows = []
    for n in GLYC_ORDER + APO_ORDER:
        base  = n.replace('_apo','')
        fname = (f"{base}_apo_APBS.csv" if '_apo' in n
                 else f"{n}_APBS_glycanAware.csv")
        fpath = CSV_DIR / fname
        if fpath.exists():
            df = pd.read_csv(fpath)
            surf = df[(df['Type']=='Protein') & (df['SurfaceLabel']=='Surface')]
            sp = ('Gallus' if n.startswith('G') else
                  'Anas'   if n.startswith('A') else 'Columba')
            apbs_rows.append({
                'Name': n, 'Species': sp,
                'State': 'Deglycosylated (Apo)' if '_apo' in n else 'Glycosylated',
                'N Surface Residues': len(surf),
                'APBS Mean (kT/e)':  round(surf['APBS_kT_e'].mean(), 4),
                'APBS Median (kT/e)': round(surf['APBS_kT_e'].median(), 4),
                'APBS Std':          round(surf['APBS_kT_e'].std(), 4),
                'APBS P5 (kT/e)':    round(surf['APBS_kT_e'].quantile(0.05), 4),
                'APBS P95 (kT/e)':   round(surf['APBS_kT_e'].quantile(0.95), 4),
            })
    apbs_df = pd.DataFrame(apbs_rows)

    with pd.ExcelWriter(out, engine='openpyxl') as writer:

        # ── Sheet A: Strip chart — raw APBS per-residue (一个子表来展示) ────
        # 用 summary 表展示代替完整残基级别数据（避免文件过大）
        strip_out = summary[['Name','Species','State',
                              'N_surface','N_hotspot','APBS_mean',
                              'APBS_median','APBS_std',
                              'APBS_P5','APBS_P95']].rename(columns={
            'Name':'Structure','N_surface':'N Surface Residues',
            'N_hotspot':'N Hotspot Residues',
            'APBS_mean':'APBS Mean (kT/e)','APBS_median':'APBS Median (kT/e)',
            'APBS_std':'APBS Std','APBS_P5':'APBS P5 (kT/e)',
            'APBS_P95':'APBS P95 (kT/e)',
        })
        strip_out.to_excel(writer, sheet_name='PanelA_APBS_Summary', index=False)
        ws = writer.sheets['PanelA_APBS_Summary']
        ws.freeze_panes = 'A2'
        for j in range(1, len(strip_out.columns)+1):
            style_header(ws.cell(row=1, column=j),
                         value=strip_out.columns[j-1])
            set_col_width(ws, j, 24)

        # ── Sheet A-full: full per-structure APBS stats ──────────────────────
        apbs_df.to_excel(writer, sheet_name='PanelA_Per-Structure_APBS',
                         index=False)
        ws2 = writer.sheets['PanelA_Per-Structure_APBS']
        ws2.freeze_panes = 'A2'
        for j in range(1, len(apbs_df.columns)+1):
            style_header(ws2.cell(row=1, column=j),
                         value=apbs_df.columns[j-1])
            set_col_width(ws2, j, 22)

        # ── Sheet B: N_hotspot glyco vs apo comparison ───────────────────────
        b_rows = []
        for sp in SPECIES_ORDER:
            glyco = summary[(summary.Species==sp) & (~summary.IsApo)]
            apo   = summary[(summary.Species==sp) & (summary.IsApo)]
            b_rows.append({
                'Species': sp,
                'Glycosylated N Structures': len(glyco),
                'Glycosylated Hotspot Count Mean': round(glyco.N_hotspot.mean(), 2),
                'Glycosylated Hotspot Count Std': round(glyco.N_hotspot.std(ddof=1), 2)
                                                  if len(glyco)>1 else 0,
                'Apo N Structures': len(apo),
                'Apo Hotspot Count Mean': round(apo.N_hotspot.mean(), 2),
                'Apo Hotspot Count Std':  round(apo.N_hotspot.std(ddof=1), 2)
                                          if len(apo)>1 else 0,
                'Δ (Glyco − Apo)': round(glyco.N_hotspot.mean() -
                                          apo.N_hotspot.mean(), 3),
            })
            if len(glyco) >= 2:
                _, p = stats.ttest_1samp(glyco.N_hotspot.values,
                                         apo.N_hotspot.values[0])
                b_rows[-1]['t-test p-value'] = round(p, 5)
                b_rows[-1]['Significance']   = ('***' if p<0.001 else '**'
                                                  if p<0.01 else '*' if p<0.05
                                                  else 'ns')
            else:
                b_rows[-1]['t-test p-value'] = 'n/a (n=1)'
                b_rows[-1]['Significance']   = 'n/a'
        dbb = pd.DataFrame(b_rows)
        dbb.to_excel(writer, sheet_name='PanelB_Hotspot_Count', index=False)
        ws = writer.sheets['PanelB_Hotspot_Count']
        ws.freeze_panes = 'A2'
        for j in range(1, len(dbb.columns)+1):
            style_header(ws.cell(row=1, column=j), value=dbb.columns[j-1])
            set_col_width(ws, j, 28)

        # ── Sheet C: Carboxylate SASA glyco vs apo ───────────────────────────
        sasa_df.to_excel(writer, sheet_name='PanelC_Carboxylate_SASA_Raw',
                         index=False)
        ws = writer.sheets['PanelC_Carboxylate_SASA_Raw']
        ws.freeze_panes = 'A2'
        for j in range(1, len(sasa_df.columns)+1):
            style_header(ws.cell(row=1, column=j), value=sasa_df.columns[j-1])
            set_col_width(ws, j, 28)

        # summary comparison
        c_rows = []
        for sp in SPECIES_ORDER:
            glyco = sasa_df[(sasa_df.Species==sp) & (sasa_df.State=='Glycosylated')]
            apo   = sasa_df[(sasa_df.Species==sp) &
                            (sasa_df.State=='Deglycosylated (Apo)')]
            col   = 'Carboxylate Surface SASA (Å²)'
            c_rows.append({
                'Species': sp,
                'Glycosylated n': len(glyco),
                'Glycosylated SASA Mean (Å²)': round(glyco[col].mean(), 1),
                'Glycosylated SASA Std':        round(glyco[col].std(ddof=1), 1)
                                                if len(glyco)>1 else 0,
                'Apo SASA (Å²)':               round(apo[col].mean(), 1),
                'ΔSASA (Glyco − Apo) (Å²)':    round(glyco[col].mean() -
                                                      apo[col].mean(), 1),
            })
            if len(glyco) >= 2:
                _, p = stats.ttest_1samp(glyco[col].values, apo[col].values[0])
                c_rows[-1]['t-test p-value'] = round(p, 5)
                c_rows[-1]['Significance']   = ('***' if p<0.001 else '**'
                                                  if p<0.01 else '*'
                                                  if p<0.05 else 'ns')
            else:
                c_rows[-1]['t-test p-value'] = 'n/a (n=1)'
                c_rows[-1]['Significance']   = 'n/a'
        dfc = pd.DataFrame(c_rows)
        dfc.to_excel(writer, sheet_name='PanelC_SASA_Comparison', index=False)
        ws = writer.sheets['PanelC_SASA_Comparison']
        ws.freeze_panes = 'A2'
        for j in range(1, len(dfc.columns)+1):
            style_header(ws.cell(row=1, column=j), value=dfc.columns[j-1])
            set_col_width(ws, j, 28)

        # ── Sheet D: APBS median violin — glyco vs apo per species ──────────
        d_rows = []
        for sp in SPECIES_ORDER:
            glyco = summary[(summary.Species==sp) & (~summary.IsApo)]
            apo   = summary[(summary.Species==sp) & (summary.IsApo)]
            d_rows.append({
                'Species': sp,
                'Glycosylated n': len(glyco),
                'Glyco APBS Median Mean (kT/e)': round(glyco.APBS_median.mean(), 4),
                'Glyco APBS Median Std':          round(glyco.APBS_median.std(ddof=1), 4)
                                                  if len(glyco)>1 else 0,
                'Apo n': len(apo),
                'Apo APBS Median Mean (kT/e)':   round(apo.APBS_median.mean(), 4),
                'Δ APBS Median (Glyco − Apo)':   round(glyco.APBS_median.mean() -
                                                        apo.APBS_median.mean(), 4),
            })
            if len(glyco) >= 2:
                _, p = stats.ttest_1samp(glyco.APBS_median.values,
                                         apo.APBS_median.values[0])
                d_rows[-1]['t-test p-value'] = round(p, 5)
                d_rows[-1]['Significance']   = ('***' if p<0.001 else '**'
                                                  if p<0.01 else '*'
                                                  if p<0.05 else 'ns')
            else:
                d_rows[-1]['t-test p-value'] = 'n/a (n=1)'
                d_rows[-1]['Significance']   = 'n/a'
        # 附上各结构 APBS median 明细
        d_detail = summary[['Name','Species','State',
                             'APBS_median','APBS_mean','APBS_std',
                             'APBS_P5','APBS_P95','N_hotspot',
                             'N_strong_neg','N_strong_pos']].rename(columns={
            'Name':'Structure','APBS_median':'APBS Median (kT/e)',
            'APBS_mean':'APBS Mean (kT/e)','APBS_std':'APBS Std',
            'APBS_P5':'APBS P5','APBS_P95':'APBS P95',
            'N_hotspot':'N Hotspot','N_strong_neg':'N Strong Neg (<-5kT/e)',
            'N_strong_pos':'N Strong Pos (>5kT/e)',
        })
        dfd = pd.DataFrame(d_rows)
        dfd.to_excel(writer, sheet_name='PanelD_APBS_Comparison', index=False)
        ws = writer.sheets['PanelD_APBS_Comparison']
        ws.freeze_panes = 'A2'
        for j in range(1, len(dfd.columns)+1):
            style_header(ws.cell(row=1, column=j), value=dfd.columns[j-1])
            set_col_width(ws, j, 28)

        d_detail.to_excel(writer, sheet_name='PanelD_Per-Structure_Detail',
                          index=False)
        ws = writer.sheets['PanelD_Per-Structure_Detail']
        ws.freeze_panes = 'A2'
        for j in range(1, len(d_detail.columns)+1):
            style_header(ws.cell(row=1, column=j), value=d_detail.columns[j-1])
            set_col_width(ws, j, 24)

    print(f"  ✓ {out.name}")


# ════════════════════════════════════════════════════════════════════════════════
# 主程序
# ════════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print("导出数据表格...")
    export_figure1()
    export_figure2()
    export_figure3()
    export_figure4()
    print("\n全部完成！输出文件位置：")
    for f in sorted(FOLDER.glob("Figure*.xlsx")):
        print(f"  {f.name}  ({f.stat().st_size // 1024} KB)")
