from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


OUTPUT_PATH = Path(r"e:\Data\Desktop\Work On\eggshell_bone_matrix_extraction_methods.xlsx")


HEADERS = [
    "来源范围",
    "文献",
    "DOI",
    "年份",
    "样本/目标",
    "脱矿试剂",
    "巯基乙醇",
    "细胞裂解液/蛋白提取液",
    "盐酸胍/胍盐",
    "其他关键试剂",
    "提取/富集说明",
    "下游检测/用途",
    "方法信息精度",
    "备注",
]


EGGSHELL_ROWS = [
    {
        "来源范围": "补充经典提取文献",
        "文献": "Panheleux M, Nys Y, Williams J, Gautron J, Boldicke T, Hincke MT. 2000. Extraction and quantification by ELISA of eggshell organic matrix proteins (ovocleidin-17, ovalbumin, ovotransferrin) in shell from young and old hens.",
        "DOI": "10.1093/ps/79.4.580",
        "年份": 2000,
        "样本/目标": "鸡蛋壳碎片中的可溶性有机基质蛋白（OC-17、ovalbumin、ovotransferrin）",
        "脱矿试剂": "20% 乙酸",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "未说明",
        "盐酸胍/胍盐": "未使用/未说明",
        "其他关键试剂": "冻干；ELISA 抗体体系",
        "提取/富集说明": "先以乙酸脱矿释放蛋壳有机基质，再进入 ELISA 定量。",
        "下游检测/用途": "ELISA 定量蛋壳基质蛋白，比较不同产蛋阶段/年龄。",
        "方法信息精度": "高：检索片段直接给出 20% 乙酸脱矿。",
        "备注": "这是当前最适合直接复现的经典蛋壳基质酸提路线之一。",
    },
    {
        "来源范围": "当前蛋壳 PTM 文献",
        "文献": "Mann K. 1999. Isolation of a glycosylated form of the chicken eggshell protein ovocleidin and determination of the glycosylation site. Alternative glycosylation/phosphorylation at an N-glycosylation sequon.",
        "DOI": "10.1016/S0014-5793(99)01586-0",
        "年份": 1999,
        "样本/目标": "蛋壳钙化层可溶性有机基质中的 ovocleidin 糖基化形式",
        "脱矿试剂": "未直接说明；来源为 soluble eggshell matrix，通常属于酸溶基质路线",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "可溶性蛋壳基质提取液",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "色谱分离缓冲体系",
        "提取/富集说明": "在可溶性基质提取物中进一步分离 glycosylated ovocleidin，再定位糖基化位点。",
        "下游检测/用途": "糖基化位点鉴定与糖基化/磷酸化替代关系分析。",
        "方法信息精度": "中：摘要能确认来自 soluble organic matrix，但未直接给出全部缓冲液配方。",
        "备注": "若做复现，通常需结合同组的乙酸酸溶提取或酸溶有机基质 protocol。",
    },
    {
        "来源范围": "当前蛋壳 PTM 文献",
        "文献": "Mann K, Hincke MT, Nys Y. 2002. Isolation of ovocleidin-116 from chicken eggshells, correction of its amino acid sequence and identification of disulfide bonds and glycosylated Asn.",
        "DOI": "10.1016/S0945-053X(02)00031-8",
        "年份": 2002,
        "样本/目标": "鸡蛋壳可溶性基质中的 ovocleidin-116 及其内源性片段",
        "脱矿试剂": "未直接说明；文摘仅说明来自 soluble chicken eggshell matrix",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "可溶性蛋壳基质提取液",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "层析分离缓冲体系",
        "提取/富集说明": "先得到可溶性蛋壳基质，再经层析分级获得 OC-116 片段并做序列/糖基化分析。",
        "下游检测/用途": "序列校正、二硫键映射、糖基化位点鉴定。",
        "方法信息精度": "中：检索片段直接说明 chromatographic fractionation，但未给出完整配方。",
        "备注": "适合把它看作“酸溶可溶性基质 + 色谱纯化 OC-116”的路线。",
    },
    {
        "来源范围": "补充经典提取文献",
        "文献": "Mann K, Macek B, Olsen JV. 2006. Proteomic analysis of the acid-soluble organic matrix of the chicken calcified eggshell layer.",
        "DOI": "10.1002/pmic.200600120",
        "年份": 2006,
        "样本/目标": "鸡蛋壳钙化层酸溶性有机基质蛋白",
        "脱矿试剂": "未直接说明；标题明确为 acid-soluble organic matrix，通常属于酸脱矿体系",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "酸溶性蛋壳基质提取液",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "蛋白酶解试剂",
        "提取/富集说明": "先获得钙化层 acid-soluble fraction，再做蛋白组鉴定。",
        "下游检测/用途": "LC-MS/MS 蛋白组鉴定酸溶性蛋壳基质蛋白。",
        "方法信息精度": "中：题名能直接确认 acid-soluble 提取目标，但浓度/时间需回原文方法节。",
        "备注": "这篇最适合补足‘蛋壳酸溶性基质’这一类蛋白组方法。",
    },
    {
        "来源范围": "补充经典提取文献",
        "文献": "Miksik I, Eckhardt A, Sedlakova P, Mikulikova K. 2007. Proteins of Insoluble Matrix of Avian (Gallus gallus) Eggshell.",
        "DOI": "10.1080/03008200601003116",
        "年份": 2007,
        "样本/目标": "水不溶/EDTA 不溶的蛋壳不溶性基质蛋白",
        "脱矿试剂": "EDTA（文摘可确认使用后得到 EDTA-insoluble fraction）",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "未说明",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "酶切试剂；HPLC-MS 流动相",
        "提取/富集说明": "先做 EDTA 脱矿，收集 EDTA-insoluble 残余基质，再酶切后 HPLC-MS 鉴定。",
        "下游检测/用途": "不溶性蛋壳基质蛋白鉴定。",
        "方法信息精度": "中高：检索片段直接给出 EDTA-insoluble 和 enzymatic cleavage。",
        "备注": "与酸溶性有机基质路线互补，适合覆盖 intramineral/insoluble 组分。",
    },
    {
        "来源范围": "当前蛋壳 PTM 文献",
        "文献": "Mann K, Olsen JV, Macek B, Gnad F, Mann M. 2007. Phosphoproteins of the chicken eggshell calcified layer.",
        "DOI": "10.1002/pmic.200600635",
        "年份": 2007,
        "样本/目标": "鸡蛋壳钙化层磷蛋白/磷肽",
        "脱矿试剂": "未说明",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "蛋壳基质蛋白提取液（具体配方未说明）",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "多种 cleavage 试剂；阴离子交换富集介质",
        "提取/富集说明": "先得到蛋壳钙化层基质蛋白，再切割并做 phosphopeptide 阴离子交换富集。",
        "下游检测/用途": "磷位点鉴定与磷蛋白组分析。",
        "方法信息精度": "中高：Scholar 片段直接给出 several cleavage methods 和 anion-exchange enrichment。",
        "备注": "这一条更偏“提取后富集”的 phosphoproteomics workflow。",
    },
    {
        "来源范围": "当前蛋壳 PTM 文献",
        "文献": "Yang R, Geng F, Huang X, Qiu N, Li S, Teng H, Chen L, Song H, Huang Q. 2020. Integrated proteomic, phosphoproteomic and N-glycoproteomic analyses of chicken eggshell matrix.",
        "DOI": "10.1016/j.foodchem.2020.127167",
        "年份": 2020,
        "样本/目标": "鸡蛋壳基质总蛋白、磷蛋白组和 N-糖蛋白组",
        "脱矿试剂": "未直接说明",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "蛋白提取缓冲液",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "胰蛋白酶；磷肽富集材料；N-糖蛋白/糖肽富集试剂",
        "提取/富集说明": "先提取蛋壳基质总蛋白，再分流至 phosphoproteomics 与 N-glycoproteomics。",
        "下游检测/用途": "整合蛋白组、磷蛋白组和 N-糖蛋白组。",
        "方法信息精度": "中：题名和组学设计可确认 workflow 类型，但具体缓冲体系需回原文方法节。",
        "备注": "适合作为现代多组学版本的蛋壳基质提取-富集流程概括。",
    },
    {
        "来源范围": "当前蛋壳 PTM 文献",
        "文献": "Zeng L, Shi X, Lin X, Zheng J. 2023. Comparative N-Glycoproteomic Investigation of Eggshell Cuticle and Mineralized Layer Proteins.",
        "DOI": "10.1021/acs.jafc.3c00708",
        "年份": 2023,
        "样本/目标": "蛋壳角质层与矿化层蛋白中的 N-糖蛋白",
        "脱矿试剂": "未直接说明",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "蛋白提取缓冲液",
        "盐酸胍/胍盐": "未说明",
        "其他关键试剂": "分层取样；胰酶；N-糖蛋白或糖肽富集试剂",
        "提取/富集说明": "分别提取 cuticle 与 mineralized layer 蛋白，再进入 N-glycoproteomics 流程。",
        "下游检测/用途": "比较角质层与矿化层 N-糖蛋白组成。",
        "方法信息精度": "中：题名和摘要能确认分层 N-glycoproteomics，但具体富集化学需查原文方法。",
        "备注": "适合研究蛋壳不同结构层的糖蛋白提取差异。",
    },
]


BONE_ROWS = [
    {
        "来源范围": "骨基质扩展",
        "文献": "Sampath TK, Reddi AH. 1984. Distribution of bone inductive proteins in mineralized and demineralized extracellular matrix.",
        "DOI": "10.1016/0006-291X(84)90865-9",
        "年份": 1984,
        "样本/目标": "矿化骨基质中的骨诱导蛋白/非胶原蛋白",
        "脱矿试剂": "0.5 M EDTA",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "未使用常规细胞裂解液",
        "盐酸胍/胍盐": "4 M guanidine-HCl；4 M guanidine-HCl + 0.5 M EDTA",
        "其他关键试剂": "顺序分级提取缓冲体系",
        "提取/富集说明": "顺序提取：先 4 M guanidine-HCl，再 EDTA 脱矿，最后 guanidine-HCl + EDTA 提取残余强结合组分。",
        "下游检测/用途": "区分骨诱导蛋白在矿化前相、脱矿相和残余相中的分布。",
        "方法信息精度": "高：Scholar 片段直接给出 sequential extraction 组成。",
        "备注": "这是骨基质经典的 sequential extraction 框架。",
    },
    {
        "来源范围": "骨基质扩展",
        "文献": "Takagi M, Maeno M, Takahashi Y, Otsuka K. 1992. Biochemical and immuno- and lectin-histochemical studies of solubility and retention of bone matrix proteins during EDTA demineralization.",
        "DOI": "10.1007/BF01082443",
        "年份": 1992,
        "样本/目标": "新鲜或固定骨样本中的骨基质蛋白",
        "脱矿试剂": "水相 EDTA",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "未使用常规细胞裂解液",
        "盐酸胍/胍盐": "4 M guanidine-HCl（前后两步都用）",
        "其他关键试剂": "免疫组化/凝集素染色相关试剂",
        "提取/富集说明": "4 M guanidine-HCl 预提，再用不含 GdnHCl 的 EDTA 脱矿，随后再次用 GdnHCl 提取残余蛋白。",
        "下游检测/用途": "评估骨基质蛋白在 EDTA 脱矿过程中的溶解性和保留情况。",
        "方法信息精度": "高：PubMed 摘要直接给出 three-step extraction。",
        "备注": "适合你把蛋壳 EDTA 不溶组分与骨中 EDTA 脱矿行为对照着看。",
    },
    {
        "来源范围": "骨基质扩展",
        "文献": "Jiang X, Ye M, Liu G, Feng S, Cui L, Zou H. 2007. Method development of efficient protein extraction in bone tissue for proteome analysis.",
        "DOI": "10.1021/pr070056t",
        "年份": 2007,
        "样本/目标": "骨组织总蛋白及水不溶性非胶原蛋白，用于蛋白组学",
        "脱矿试剂": "HCl；0.5 M tetrasodium EDTA（pH 7.4）",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "未使用常规细胞裂解液",
        "盐酸胍/胍盐": "guanidine-HCl",
        "其他关键试剂": "蛋白组学样品制备试剂",
        "提取/富集说明": "先用 HCl 脱矿提高提取效率，再结合 guanidine-HCl 和 EDTA 顺序提取骨蛋白。",
        "下游检测/用途": "骨组织 proteome analysis。",
        "方法信息精度": "高：Scholar 片段直接给出 HCl 脱矿、guanidine-HCl 与 EDTA。",
        "备注": "如果目标是蛋白组学而不是单一蛋白纯化，这条最实用。",
    },
    {
        "来源范围": "骨基质扩展",
        "文献": "Cleland TP, Voegele K, Schweitzer MH. 2012. Empirical Evaluation of Bone Extraction Protocols.",
        "DOI": "10.1371/journal.pone.0031443",
        "年份": 2012,
        "样本/目标": "现代和古骨中的 collagen I、osteocalcin、非胶原蛋白",
        "脱矿试剂": "HCl；EDTA（两类方案均被直接比较）",
        "巯基乙醇": "未说明",
        "细胞裂解液/蛋白提取液": "未使用常规细胞裂解液",
        "盐酸胍/胍盐": "guanidine-HCl",
        "其他关键试剂": "ammonium bicarbonate；SDS；urea/thiourea",
        "提取/富集说明": "系统比较 HCl 与 EDTA 脱矿，以及 guanidine-HCl、ammonium bicarbonate 等后续溶出体系。",
        "下游检测/用途": "SDS-PAGE、ELISA、古蛋白组/现代骨蛋白提取方法比较。",
        "方法信息精度": "高：PLOS 正文直接比较不同脱矿与溶出缓冲体系。",
        "备注": "若你要把骨提取路线迁移到壳基质比较，最有价值的是其‘纯度 vs. 兼容质谱’判断。",
    },
]


def write_sheet(worksheet, title: str, rows: list[dict[str, object]]) -> None:
    worksheet.title = title
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append([row[header] for header in HEADERS])

    worksheet.freeze_panes = "A2"
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    widths = {
        "A": 18,
        "B": 68,
        "C": 26,
        "D": 10,
        "E": 34,
        "F": 24,
        "G": 14,
        "H": 24,
        "I": 24,
        "J": 28,
        "K": 36,
        "L": 30,
        "M": 18,
        "N": 28,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def build_workbook() -> Workbook:
    workbook = Workbook()
    eggshell_sheet = workbook.active
    write_sheet(eggshell_sheet, "蛋壳基质提取", EGGSHELL_ROWS)
    bone_sheet = workbook.create_sheet("骨基质提取")
    write_sheet(bone_sheet, "骨基质提取", BONE_ROWS)
    return workbook


def main() -> None:
    workbook = build_workbook()
    workbook.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")
    print(f"Eggshell rows: {len(EGGSHELL_ROWS)}")
    print(f"Bone rows: {len(BONE_ROWS)}")


if __name__ == "__main__":
    main()